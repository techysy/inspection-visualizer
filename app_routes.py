import io
import re
import base64
import tempfile
import logging
import logging.handlers
from datetime import datetime, date, timedelta
from pathlib import Path
import json
from flask import Blueprint, render_template, jsonify, request, redirect, url_for, flash, Response, send_file, session
from werkzeug.security import generate_password_hash, check_password_hash
from PIL import Image
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from models.inspection import SessionLocal, InspectionObject, InspectionRecord, Inspector, ObjectMetric, DailyListRecord

log_dir = Path(__file__).parent / 'log'
log_dir.mkdir(exist_ok=True)
log_file = log_dir / 'app.log'
log_handler = logging.handlers.TimedRotatingFileHandler(
    log_file, when='midnight', interval=1, backupCount=30, encoding='utf-8'
)
log_handler.suffix = '%Y-%m-%d'
log_handler.namer = lambda name: str(log_dir / f'app{name[-10:]}.log')

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
    handlers=[
        log_handler,
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)


def get_image_creation_time(file_storage):
    """从图片EXIF获取拍摄时间，失败则返回None"""
    try:
        file_storage.seek(0)
        img = Image.open(file_storage)
        exif_data = img._getexif()
        if exif_data:
            # EXIF tag 36867 = DateTimeOriginal
            date_taken = exif_data.get(36867)
            if date_taken:
                dt = datetime.strptime(date_taken, '%Y:%m:%d %H:%M:%S')
                logger.info(f'  IMAGE: EXIF拍摄时间 {dt}')
                file_storage.seek(0)
                return dt
            # EXIF tag 306 = DateTime
            date_modified = exif_data.get(306)
            if date_modified:
                dt = datetime.strptime(date_modified, '%Y:%m:%d %H:%M:%S')
                logger.info(f'  IMAGE: EXIF修改时间 {dt}')
                file_storage.seek(0)
                return dt
    except Exception as e:
        logger.info(f'  IMAGE: EXIF读取失败 {e}')
    file_storage.seek(0)
    return None
logger = logging.getLogger(__name__)

OCR_CONFIG_PATH = Path(__file__).parent / 'ocr_config.json'
DASHBOARD_TYPES_PATH = Path(__file__).parent / 'dashboard_types.json'

_default_ocr_config = {
    "text_score": 0.5,
    "use_det": True,
    "use_cls": True,
    "use_rec": True,
    "min_height": 30,
    "max_side_len": 2000,
    "ignore_top": 0,
    "ignore_bottom": 0,
}


def _load_ocr_config():
    if OCR_CONFIG_PATH.exists():
        try:
            with open(OCR_CONFIG_PATH, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception:
            pass
    return _default_ocr_config.copy()


def _build_ocr_engine(config=None):
    if config is None:
        config = _load_ocr_config()
    try:
        from rapidocr_onnxruntime import RapidOCR
        kwargs = {}
        if 'text_score' in config:
            kwargs['text_score'] = config['text_score']
        if 'use_det' in config:
            kwargs['use_det'] = config['use_det']
        if 'use_cls' in config:
            kwargs['use_cls'] = config['use_cls']
        if 'use_rec' in config:
            kwargs['use_rec'] = config['use_rec']
        return RapidOCR(**kwargs)
    except ImportError:
        return None


ocr_engine = _build_ocr_engine()


def _load_dashboard_types():
    if DASHBOARD_TYPES_PATH.exists():
        try:
            with open(DASHBOARD_TYPES_PATH, 'r', encoding='utf-8') as f:
                data = json.load(f)
                return data.get('types', [])
        except Exception:
            pass
    return []


def _save_dashboard_types(types_list):
    import tempfile
    tmp_path = DASHBOARD_TYPES_PATH.with_suffix('.tmp')
    with open(tmp_path, 'w', encoding='utf-8') as f:
        json.dump({'types': types_list}, f, ensure_ascii=False, indent=2)
    tmp_path.replace(DASHBOARD_TYPES_PATH)

def _get_virtual_metric_keys():
    """返回所有 dashboard type calc_config 中定义的 result_name 集合（虚拟指标不可删除）"""
    keys = set()
    for t in _load_dashboard_types():
        calc_configs = t.get('calc_configs') or []
        single_config = t.get('calc_config')
        if single_config:
            calc_configs = [single_config] + calc_configs
        for cfg in calc_configs:
            if cfg.get('result_name'):
                keys.add(cfg['result_name'])
    return keys

def _get_virtual_metric_keys_for_object(obj):
    """返回指定巡检对象所属类型的虚拟指标 key 集合"""
    keys = set()
    for dt in _load_dashboard_types():
        if dt.get('category') == obj.device_type:
            keys.update(_get_virtual_metric_keys_for_type(dt))
    return keys

def _get_virtual_metric_keys_for_type(dtype):
    """返回指定 dashboard type 的 calc_config.result_name（仅该类型的虚拟指标）"""
    keys = []
    # 支持单个 calc_config 或数组 calc_configs
    calc_configs = dtype.get('calc_configs') or []
    single_config = dtype.get('calc_config')
    if single_config:
        calc_configs = [single_config] + calc_configs
    for cfg in calc_configs:
        if cfg.get('result_name'):
            keys.append(cfg['result_name'])
    return keys


def _sync_virtual_metrics_for_dtype(dtype, session):
    """同步仪表盘类型的虚拟指标到巡检对象"""
    calc_configs = dtype.get('calc_configs') or []
    single_config = dtype.get('calc_config')
    if single_config:
        calc_configs = [single_config] + calc_configs
    if not calc_configs:
        return
    category = dtype.get('category', '')
    if not category:
        return

    objects = session.query(InspectionObject).filter(
        InspectionObject.status == 'active',
        ((InspectionObject.device_type == category) | InspectionObject.device_type.is_(None))
    ).all()
    for calc in calc_configs:
        if not calc or not calc.get('result_name'):
            continue
        result_name = calc.get('result_name')
        unit = calc.get('unit', '')
        for obj in objects:
            existing = session.query(ObjectMetric).filter_by(object_id=obj.id, key=result_name).first()
            if existing:
                existing.show_in_chart = calc.get('show_chart', False)
                existing.max_value = calc.get('max_value')
                if unit:
                    existing.unit = unit
            else:
                session.add(ObjectMetric(
                    object_id=obj.id, key=result_name, name=result_name, unit=unit,
                    max_value=calc.get('max_value') or 100,
                    show_in_chart=calc.get('show_chart', False)
                ))


def _sync_dashboard_type_from_object(obj, metrics):
    """根据巡检对象及其指标自动创建/更新仪表盘类型"""
    types_list = _load_dashboard_types()
    type_id = f'obj_{obj.id}'

    existing = next((t for t in types_list if t['id'] == type_id), None)

    labels = {}
    for m in metrics:
        labels[m.name] = m.key

    type_data = {
        'id': type_id,
        'name': obj.name,
        'description': obj.description or f'{obj.name} 仪表盘',
        'detect_keywords': [obj.name],
        'labels': labels,
        'extra_labels': [],
        'result_rules': {},
        'number_before_label': False,
    }

    if existing:
        existing.update(type_data)
    else:
        types_list.append(type_data)

    _save_dashboard_types(types_list)
    return type_data


def _sync_bidirectional(obj_id, dashboard_type_id, session):
    """双向同步指标：仪表盘类型↔所有关联对象"""
    if not dashboard_type_id:
        return

    # 加载仪表盘类型
    types_list = _load_dashboard_types()
    matched_dtype = next((dt for dt in types_list if dt.get('id') == dashboard_type_id), None)
    if not matched_dtype:
        return

    # 通过名称/category匹配关联对象（模型无 dashboard_type_id 字段）
    try:
        all_objects = session.query(InspectionObject).filter_by(status='active').all()
    except Exception:
        return
    dtype_name = matched_dtype.get('name', '')
    dtype_category = matched_dtype.get('category', '')
    related_objects = []
    for obj in all_objects:
        if obj.name == dtype_name or obj.device_type == dtype_category:
            related_objects.append(obj)
    # 确保当前对象在列表中
    if obj_id not in [o.id for o in related_objects]:
        obj = session.query(InspectionObject).get(obj_id)
        if obj:
            related_objects.append(obj)

    if not related_objects:
        return

    dtype_labels = matched_dtype.get('labels', {})

    # 方向1：仪表盘类型 → 所有关联对象
    if dtype_labels:
        for obj in related_objects:
            obj_metrics = session.query(ObjectMetric).filter_by(object_id=obj.id).all()
            obj_keys = {m.key: m for m in obj_metrics}
            obj_names = {m.name: m for m in obj_metrics}
            for cn_label, short_key in dtype_labels.items():
                if short_key not in obj_keys and cn_label not in obj_names:
                    session.add(ObjectMetric(
                        object_id=obj.id, key=short_key, name=cn_label,
                        sort_order=len(obj_metrics)
                    ))
                    logger.info(f'  SYNC: 仪表盘→对象({obj.id}) 创建 {cn_label}({short_key})')

    # 方向2：对象 → 仪表盘类型（类型无标签时，从有指标的对象同步）
    if not dtype_labels:
        for obj in related_objects:
            obj_metrics = session.query(ObjectMetric).filter_by(object_id=obj.id).all()
            if obj_metrics:
                for m in obj_metrics:
                    if m.name not in dtype_labels:
                        dtype_labels[m.name] = m.key
                        logger.info(f'  SYNC: 对象({obj.id})→仪表盘 添加 {m.name}({m.key})')
                break  # 只从第一个有指标的对象同步

    # 方向2补充：对象有指标但类型没有的，也添加到类型
    for obj in related_objects:
        obj_metrics = session.query(ObjectMetric).filter_by(object_id=obj.id).all()
        for m in obj_metrics:
            if m.name not in dtype_labels:
                dtype_labels[m.name] = m.key
                logger.info(f'  SYNC: 对象({obj.id})→仪表盘 补充 {m.name}({m.key})')

    # 保存仪表盘类型更新
    matched_dtype['labels'] = dtype_labels
    _save_dashboard_types(types_list)


main = Blueprint('main', __name__)


# ──────────────────────── 鉴权 ────────────────────────

PUBLIC_ROUTES = {'login', 'api_login', 'api_auth_status', 'static'}
ADMIN_ROUTES = {'inspectors', 'inspector_add', 'inspector_edit', 'inspector_delete', 'ocr_admin', 'bulk_import', 'objects', 'object_add', 'object_edit', 'object_delete', 'object_clone', 'backup_gallery_page', 'api_backup_gallery', 'serve_backup_image'}

def _check_request_auth():
    """检查请求是否已认证（session 或请求体中的 username+password）"""
    if session.get('logged_in'):
        return True
    if request.is_json:
        data = request.get_json(silent=True) or {}
        uname = data.get('username', '').strip()
        pw = data.get('password', '')
        if uname and pw:
            session_local = SessionLocal()
            try:
                inspector = session_local.query(Inspector).filter(
                    Inspector.username == uname
                ).first()
                if inspector and inspector.password and check_password_hash(inspector.password, pw):
                    session['logged_in'] = True
                    session['inspector_id'] = inspector.id
                    session['inspector_name'] = inspector.name
                    session['is_admin'] = inspector.is_admin or False
                    session.permanent = True
                    return True
            finally:
                session_local.close()
        # 兼容旧方式：仅 password（无 username）时使用 APP_PASSWORD
        if not uname and pw:
            from flask import current_app
            if pw == current_app.config.get('APP_PASSWORD', ''):
                session['logged_in'] = True
                session['inspector_id'] = None
                session['inspector_name'] = ''
                session['is_admin'] = True  # APP_PASSWORD 视为管理员
                session.permanent = True
                return True
    return False


def _require_inspector():
    """当前登录的巡检人员信息"""
    return {
        'id': session.get('inspector_id'),
        'name': session.get('inspector_name', ''),
    }


@main.before_request
def check_auth():
    """所有路由需要登录（除公开路由外）"""
    ep = request.endpoint or ''
    if not ep.startswith('main.'):
        return None
    route = ep.split('main.', 1)[1]
    if route in PUBLIC_ROUTES:
        return None
    if _check_request_auth():
        # 管理员路由检查
        if route in ADMIN_ROUTES and not session.get('is_admin'):
            if request.path.startswith('/api/'):
                return jsonify({'error': '无权限，需要管理员身份'}), 403
            return redirect(url_for('main.index'))
        return None
    if request.path.startswith('/api/'):
        return jsonify({'error': '未登录', 'login_url': url_for('main.login', _external=True)}), 401
    return redirect(url_for('main.login'))


@main.route('/login', methods=['GET'])
def login():
    """登录页面"""
    if session.get('logged_in'):
        return redirect(url_for('main.index'))
    return render_template('login.html')


@main.route('/api/login', methods=['POST'])
def api_login():
    """API 登录：使用巡检人员用户名+密码"""
    data = request.get_json()
    if not data:
        return jsonify({'error': '请提供登录信息'}), 400
    uname = data.get('username', '').strip()
    pw = data.get('password', '')

    # 无用户名时尝试旧密码（兼容）
    if not uname:
        from flask import current_app
        if pw == current_app.config.get('APP_PASSWORD', ''):
            session['logged_in'] = True
            session['inspector_id'] = None
            session['inspector_name'] = ''
            session['is_admin'] = True
            session.permanent = True
            return jsonify({'ok': True, 'inspector': None})
        return jsonify({'error': '密码错误'}), 403

    if not pw:
        return jsonify({'error': '请输入密码'}), 400

    session_local = SessionLocal()
    try:
        inspector = session_local.query(Inspector).filter(
            Inspector.username == uname
        ).first()
        if inspector and inspector.password and check_password_hash(inspector.password, pw):
            session['logged_in'] = True
            session['inspector_id'] = inspector.id
            session['inspector_name'] = inspector.name
            session['is_admin'] = inspector.is_admin or False
            session.permanent = True
            return jsonify({'ok': True, 'inspector': {'id': inspector.id, 'name': inspector.name, 'is_admin': inspector.is_admin or False}})
        # 无密码时视为未启用登录功能
        if inspector and not inspector.password:
            return jsonify({'error': '该人员未设置登录密码，请先联系管理员'}), 403
        return jsonify({'error': '用户名或密码错误'}), 403
    finally:
        session_local.close()


@main.route('/api/auth-status', methods=['GET'])
def api_auth_status():
    """检查登录状态"""
    return jsonify({
        'logged_in': session.get('logged_in', False),
        'is_admin': session.get('is_admin', False),
        'inspector': {
            'id': session.get('inspector_id'),
            'name': session.get('inspector_name', ''),
            'is_admin': session.get('is_admin', False),
        } if session.get('inspector_id') else None
    })


@main.route('/logout')
def logout():
    """登出"""
    session.clear()
    return redirect(url_for('main.login'))


# 设备类型关键词
DEVICE_TYPES = {
    '服务器': ['服务器', 'server', 'srv', '主机'],
    '网络设备': ['交换机', '路由器', '防火墙', 'switch', 'router', 'firewall', '网络设备'],
    '存储': ['存储', 'storage', '磁盘阵列', 'nas', 'san'],
    'UPS': ['ups', '不间断电源', '电源'],
    '空调': ['空调', '精密空调', '制冷'],
    '监控': ['监控', '摄像头', 'camera', 'cctv'],
}

# 巡检结果关键词
RESULT_KEYWORDS = {
    '正常': ['正常', '良好', 'ok', 'good', '合格', 'pass'],
    '异常': ['异常', '故障', 'error', 'fail', '不合格', '告警', '报警'],
    '需关注': ['需关注', '注意', 'warning', 'warn', '待处理'],
}


def _parse_status_to_metrics(status_detail):
    """将 status_detail 字符串解析为结构化指标 dict"""
    metrics = {}
    if not status_detail:
        return metrics
    for part in re.split(r'[;；]\s*', status_detail):
        m = re.match(r'^(.+?):\s*(.+)$', part.strip())
        if m:
            key = m.group(1).strip()
            val = m.group(2).strip()
            metrics[key] = val
    return metrics


def _check_metrics_thresholds(metrics, object_id, session):
    """根据指标阈值配置判断巡检结果，优先用 ObjectMetric 阈值，无则回退到 dashboard_types.json result_rules"""
    metric_configs = session.query(ObjectMetric).filter_by(object_id=object_id).all()
    
    worst_result = '正常'
    has_threshold = False
    
    for mc in metric_configs:
        val_str = metrics.get(mc.name) or metrics.get(mc.key)
        if val_str is None:
            continue
        val_match = re.search(r'([\d.]+)', str(val_str))
        if not val_match:
            continue
        val = float(val_match.group(1))
        
        err_thr = mc.error_threshold
        warn_thr = mc.warn_threshold
        direction = mc.threshold_direction or 'lt'
        
        # 若 ObjectMetric 未配置阈值，从 dashboard_types.json result_rules 回退
        if err_thr is None and warn_thr is None:
            obj = session.query(InspectionObject).get(object_id)
            if obj and obj.device_type:
                dashboard_types = _load_dashboard_types()
                dt = next((d for d in dashboard_types if d.get('category') == obj.device_type), None)
                if dt:
                    result_rules = dt.get('result_rules', {})
                    mc_key = mc.key
                    # 从 result_rules 中提取该指标的阈值（如 online_rate<97=异常, online_rate<100=需关注）
                    # 支持中英文 key 匹配（在线率 ↔ online_rate）
                    _key_aliases = {'在线率': 'online_rate', '在线': 'online', '离线': 'offline'}
                    _reverse_aliases = {v: k for k, v in _key_aliases.items()}
                    match_keys = {mc_key, _key_aliases.get(mc_key, ''), _reverse_aliases.get(mc_key, '')}
                    for rule, outcome in result_rules.items():
                        rm = re.match(rf'^(\w+)([><=]+)([\d.]+)$', rule)
                        if not rm or rm.group(1) not in match_keys:
                            continue
                        op, thr = rm.group(2), float(rm.group(3))
                        if outcome == '异常' and err_thr is None:
                            err_thr = thr
                            if op in ('<', '<='):
                                direction = 'lt'
                            elif op in ('>', '>='):
                                direction = 'gt'
                        elif outcome == '需关注' and warn_thr is None:
                            warn_thr = thr
                            if op in ('<', '<='):
                                direction = 'lt'
                            elif op in ('>', '>='):
                                direction = 'gt'
        
        if err_thr is None and warn_thr is None:
            continue
        
        has_threshold = True
        is_gt = direction == 'gt'
        
        if err_thr is not None:
            if (is_gt and val > err_thr) or (not is_gt and val < err_thr):
                return '异常'
        
        if warn_thr is not None:
            if (is_gt and val > warn_thr) or (not is_gt and val < warn_thr):
                worst_result = '需关注'
    
    return worst_result if has_threshold else None


def _evaluate_delta_rules(metrics, object_id, result_rules, session):
    """根据日对比规则判断巡检结果（delta 操作符）"""
    if not result_rules:
        return None
    
    # 获取前一天的记录
    today = date.today()
    yesterday = today - timedelta(days=1)
    
    # 优先从 daily_list_records 获取前一天数据
    prev_daily = session.query(DailyListRecord).filter(
        DailyListRecord.object_id == object_id,
        DailyListRecord.date < today
    ).order_by(DailyListRecord.date.desc()).first()
    
    # 也检查 inspection_records
    prev_record = session.query(InspectionRecord).filter(
        InspectionRecord.object_id == object_id,
        InspectionRecord.timestamp < datetime.now()
    ).order_by(InspectionRecord.timestamp.desc()).first()
    
    # 获取前一天的值
    prev_metrics = {}
    if prev_daily:
        try:
            prev_metrics = json.loads(prev_daily.items) if prev_daily.items else {}
            prev_metrics['当日人数'] = prev_daily.count
        except:
            pass
    if prev_record and prev_record.metrics:
        try:
            prev_metrics.update(json.loads(prev_record.metrics))
        except:
            pass
    
    if not prev_metrics:
        logger.info(f'  DELTA: 未找到前一天记录，跳过 delta 规则评估')
        return None
    
    worst_result = None
    for rule, outcome in result_rules.items():
        # delta 规则：与前一天对比
        m = re.match(r'^(\w+)_?delta([><=]+)([\d.]+)$', rule)
        if m:
            metric_key, op, threshold = m.group(1), m.group(2), float(m.group(3))
            curr_val = metrics.get(metric_key, 0)
            prev_val = prev_metrics.get(metric_key, 0)
            try:
                curr_val = float(curr_val)
                prev_val = float(prev_val)
            except (ValueError, TypeError):
                continue
            delta = curr_val - prev_val
            logger.info(f'  DELTA: {metric_key} curr={curr_val} prev={prev_val} delta={delta}')
            if _evaluate_comparison(delta, op, threshold):
                worst_result = outcome
                logger.info(f'  DELTA: 规则 {rule} -> {outcome}')
    
    return worst_result


def _extract_location_from_lines(lines):
    """从行列表中提取位置信息"""
    gvars = _load_global_vars()
    skip_words = gvars.get('skip_words', _default_global_vars['skip_words'])
    
    # 优先从前面的行中提取位置（左上角区域）
    for i, line in enumerate(lines[:5]):  # 只检查前5行
        line = line.strip()
        if not line or len(line) < 2:
            continue
        if any(sw in line for sw in skip_words):
            continue
        if re.match(r'^\d+$', line):  # 纯数字
            continue
        if re.match(r'^\d{1,3}\.\d{1,2}%$', line):  # 百分比
            continue
        # 可能是区域名称（放宽正则，允许空格和常见标点）
        if re.match(r'^[\u4e00-\u9fa5a-zA-Z0-9\-_\s,，。、]{2,30}$', line):
            return line
    
    # 如果前面没有找到，尝试所有行
    for line in lines:
        line = line.strip()
        if not line or len(line) < 2:
            continue
        if any(sw in line for sw in skip_words):
            continue
        if re.match(r'^\d+$', line):  # 纯数字
            continue
        if re.match(r'^\d{1,3}\.\d{1,2}%$', line):  # 百分比
            continue
        # 可能是区域名称（放宽正则）
        if re.match(r'^[\u4e00-\u9fa5a-zA-Z0-9\-_\s,，。、]{2,30}$', line):
            return line
    
    return ''


GLOBAL_VARS_PATH = Path(__file__).parent / 'global_vars.json'

_default_global_vars = {
    'skip_words': [
        '监控点总数', '在线', '离线', '未检测', '监控点在线率', '默认区域设置', 'admin',
        '车辆总数', '在线车辆', '今日上线', '报警车辆', '全部车辆', '全部状态',
        '工牌', '全部', '未用', '请输入', '总数', '行驶', '停车', '告警',
        '在线率', '离线率', '未检测率'
    ],
    'location_aliases': {}
}


def _load_global_vars():
    if GLOBAL_VARS_PATH.exists():
        try:
            with open(GLOBAL_VARS_PATH, 'r', encoding='utf-8') as f:
                data = json.load(f)
            merged = dict(_default_global_vars)
            # 合并 skip_words：合并默认值和已保存的值，去重
            if 'skip_words' in data:
                merged['skip_words'] = list(dict.fromkeys(_default_global_vars['skip_words'] + data['skip_words']))
            else:
                merged['skip_words'] = _default_global_vars['skip_words']
            if 'location_aliases' in data:
                merged['location_aliases'] = {**_default_global_vars.get('location_aliases', {}), **data['location_aliases']}
            return merged
        except Exception:
            pass
    return dict(_default_global_vars)


def _save_global_vars(data):
    with open(GLOBAL_VARS_PATH, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def _evaluate_comparison(val, op, threshold):
    """通用比较函数，支持 >, <, >=, <=, == 运算符"""
    if op == '>' and val > threshold:
        return True
    elif op == '<' and val < threshold:
        return True
    elif op == '>=' and val >= threshold:
        return True
    elif op == '<=' and val <= threshold:
        return True
    elif op == '==' and val == threshold:
        return True
    return False


def parse_inspection_form(ocr_result, filename=None):
    """解析巡检表单/仪表盘 OCR 结果"""
    if not ocr_result:
        return []

    lines = []
    for item in ocr_result:
        if item and len(item) >= 2:
            text = item[1]
            lines.append(text)

    all_text = '\n'.join(lines)

    # OCR 原始输出日志
    logger.info('=== OCR LINES ===')
    for i, line in enumerate(lines):
        logger.info(f'  [{i}] {repr(line)}')
    logger.info('=== END ===')

    # 尝试解析仪表盘格式（监控系统截图）
    dashboard_result = parse_dashboard_screenshot(all_text, lines, filename=filename)
    if dashboard_result:
        return [dashboard_result]

    # 传统表单格式解析
    return parse_traditional_form(lines)


def parse_dashboard_screenshot(all_text, lines, filename=None):
    """解析仪表盘截图（使用可配置的仪表盘类型规则）"""
    logger.info(f'  PARSE: all_text={repr(all_text)}')
    if filename:
        logger.info(f'  PARSE: filename={repr(filename)}')

    dashboard_types = _load_dashboard_types()
    matched_type = None

    # 0. 优先用文件名匹配（解决小字 OCR 识别不到的问题）
    if filename:
        for dtype in dashboard_types:
            keywords = dtype.get('detect_keywords', [])
            if any(kw in filename for kw in keywords):
                matched_type = dtype
                logger.info(f'  PARSE: 文件名匹配到 [{keywords}] -> {dtype["name"]}')
                break

    # 1. 没匹配到才走 OCR 文字匹配
    if not matched_type:
        for dtype in dashboard_types:
            keywords = dtype.get('detect_keywords', [])
            if any(kw in all_text for kw in keywords):
                matched_type = dtype
                logger.info(f'  PARSE: 检测到{dtype["name"]}仪表盘 (关键词: {keywords})')
                break

    if not matched_type:
        logger.info('  PARSE: 未匹配到任何仪表盘类型')
        return None

    label_map = matched_type.get('labels', {})
    number_before_label = matched_type.get('number_before_label', False)
    region_name = ''
    online_rate = ''
    metrics = {}

    # 逐行匹配（使用队列处理连续标签无数字的情况）
    name_from_first_line = matched_type.get('name_from_first_line', False)
    pending_labels = []
    unmatched_lines = [] if name_from_first_line else None
    for line in lines:
        line = line.strip()
        if not line:
            continue

        # 1) 纯整数行或括号数字行 → 归给第一个等待数字的标签
        if pending_labels:
            num_match = re.match(r'^[\(（]?(\d+)[\)）]?$', line)
            if num_match:
                num = int(num_match.group(1))
                label = pending_labels.pop(0)
                metrics[label] = num
                logger.info(f'  PARSE: 数字 {num} -> {label}')
                continue

        # 2) 匹配 "位置名（X/Y）" 或 "X/Y" 格式，自动映射到配置的标签
        m = re.match(r'^(.*?)\s*[（(]?\s*(\d+)\s*/\s*(\d+)\s*[）)]?\s*$', line)
        if m:
            region_name = m.group(1).strip()
            val1 = int(m.group(2))
            val2 = int(m.group(3))
            # 根据配置的标签映射（前两个标签分别对应 X 和 Y）
            label_keys = list(label_map.keys())
            if len(label_keys) >= 2:
                metrics[label_map[label_keys[0]]] = val1
                metrics[label_map[label_keys[1]]] = val2
                logger.info(f'  PARSE: X/Y格式 {region_name or "无位置"} {label_keys[0]}={val1} {label_keys[1]}={val2}')
            else:
                metrics['online'] = val1
                metrics['total'] = val2
                logger.info(f'  PARSE: X/Y格式 {region_name or "无位置"} 在线={val1} 总数={val2}')
            pending_labels = []
            continue

        # 2.5) 匹配 "标签: X单位/Y单位" 格式（如 "已使用/总容量: 269650GiB/407416GiB"）
        m = re.match(r'^(.*?)[:：]?\s*(\d[\d.]*)\s*[a-zA-Z]*\s*/\s*(\d[\d.]*)\s*[a-zA-Z]+', line)
        if m:
            label_text = m.group(1).strip()
            val1 = float(m.group(2))
            val2 = float(m.group(3))
            if val1 == int(val1):
                val1 = int(val1)
            if val2 == int(val2):
                val2 = int(val2)
            label_keys = list(label_map.keys())
            if len(label_keys) >= 2:
                metrics[label_map[label_keys[0]]] = val1
                metrics[label_map[label_keys[1]]] = val2
                logger.info(f'  PARSE: X单位/Y单位格式 {label_text or "无标签"} {label_keys[0]}={val1} {label_keys[1]}={val2}')
            pending_labels = []
            continue

        # 3) 匹配标签行（在行内查找所有出现的标签-数值对）
        sorted_labels = sorted(label_map.keys(), key=len, reverse=True)
        line_matched = False
        for label in sorted_labels:
            # 构建搜索模式：标签后跟可选分隔符 + 数值 + 可选单位
            # 支持格式：标签 数字、标签：数字、标签(数字)、标签（数字）
            sep = r'\s*[(\（]?\s*[·.:：]?\s*'
            end = r'\s*[)\）]?'
            search_pattern = rf'{re.escape(label)}{sep}([\d.]+)(%?){end}\s*([万亿]?)'
            if number_before_label:
                search_pattern = rf'([\d.]+)(%?)\s*([万亿]?)\s*{sep}{re.escape(label)}{end}'
            for m in re.finditer(search_pattern, line):
                matched_label = label_map[label]
                raw = m.group(1)
                pct = m.group(2) if len(m.groups()) >= 2 else ''
                unit = m.group(3) if len(m.groups()) >= 3 else ''
                if unit == '万':
                    matched_val = int(float(raw) * 10000)
                elif unit == '亿':
                    matched_val = int(float(raw) * 100000000)
                elif '.' in raw:
                    matched_val = float(raw)
                else:
                    matched_val = int(float(raw))
                if pct == '%':
                    matched_val = str(matched_val) + '%'
                metrics[matched_label] = matched_val
                logger.info(f'  PARSE: {matched_label}={matched_val}')

        if line_matched:
            pending_labels = []
            continue

        # 如果行本身就是一个标签名（没有数字），加入待处理队列等待下一行的数字
        if not line_matched:
            found_new_label = False
            for label in sorted_labels:
                if line.strip() == label:
                    pending_labels = [label]
                    found_new_label = True
                    logger.info(f'  PARSE: 标签行 "{label}" 等待数字')
                    break
            # 非标签非数字行（如"监控点总数"）不清除 pending_labels，保留等待后续数字
            if not found_new_label and unmatched_lines is not None and len(line) >= 2 and not re.match(r'^\d+$', line):
                unmatched_lines.append(line)

    # 识别在线率百分比（如果有 extra_labels 包含"在线率"）
    extra_labels = matched_type.get('extra_labels', [])
    if '在线率' in extra_labels:
        rate_match = re.search(r'(\d{1,3}\.\d{1,2})\s*%', all_text)
        if rate_match:
            online_rate = rate_match.group(0)

    # 交叉校验：用在线率修正OCR错位的数值
    if online_rate:
        rate_val = float(online_rate.replace('%', ''))
        known_total = metrics.get('total', 0)
        known_online = metrics.get('online', 0)
        known_offline = metrics.get('offline', 0)
        known_undetected = metrics.get('undetected', 0)
        if known_total > 0:
            expected_online = round(known_total * rate_val / 100)
            expected_offline = max(0, known_total - expected_online - known_undetected)
            logger.info(f'  PARSE: 交叉校验 online {known_online}->{expected_online}, offline {known_offline}->{expected_offline}')
            metrics['online'] = expected_online
            metrics['offline'] = expected_offline
        elif known_online > 0 and rate_val == 100.0 and known_offline > 0:
            logger.info(f'  PARSE: 在线率100%, 修正 offline {known_offline}->0')
            metrics['offline'] = 0

    # 计算型指标：根据公式计算
    formulas = matched_type.get('formulas', {})
    calc_config = matched_type.get('calc_config', None)
    if formulas:
        for target_key, formula in formulas.items():
            try:
                # 支持公式为字符串或对象格式 {expr, decimal_places, suffix}
                if isinstance(formula, dict):
                    formula_expr = formula.get('expr', '')
                    decimal_places = formula.get('decimal_places', 2)
                    suffix = formula.get('suffix', '')
                else:
                    formula_expr = formula
                    decimal_places = None
                    suffix = ''
                # 替换公式中的key名为对应的值（按长度倒序，避免短名误替换长名）
                eval_expr = formula_expr
                sorted_keys = sorted(label_map.items(), key=lambda x: len(x[1]), reverse=True)
                for cn_label, short_key in sorted_keys:
                    val = metrics.get(short_key, 0)
                    eval_expr = eval_expr.replace(short_key, str(val))
                # 支持 + - * / 运算
                result = eval(eval_expr)
                
                # 应用格式化（round + suffix）
                if decimal_places is not None:
                    result = round(float(result), decimal_places)
                    if suffix:
                        metrics[target_key] = f'{result}{suffix}'
                    else:
                        metrics[target_key] = result
                # 应用 calc_config 格式化（仅对百分比类型）
                elif calc_config and calc_config.get('type') == 'percentage' and target_key == calc_config.get('result_name'):
                    decimal_places = calc_config.get('decimal_places', 2)
                    fmt = calc_config.get('format', 'decimal')
                    result = round(float(result), decimal_places)
                    if fmt == 'percent':
                        # 百分比格式：公式为 已用容量/总容量，乘以100并加%
                        metrics[target_key] = f'{result * 100:.{decimal_places}f}%'
                    else:
                        metrics[target_key] = f'{result}%'
                elif isinstance(result, float) and result == int(result):
                    result = int(result)
                    metrics[target_key] = result
                else:
                    metrics[target_key] = result
                    
                logger.info(f'  PARSE: 计算指标 {target_key} = {formula} = {result}')
            except Exception as e:
                logger.info(f'  PARSE: 公式计算失败 {target_key}={formula}: {e}')

    # 自定义公式计算（基于 calc_config）
    if calc_config and calc_config.get('type') == 'custom':
        try:
            fields = calc_config.get('fields', [])
            formula = calc_config.get('formula', '')
            result_name = calc_config.get('result_name', '')
            if formula and fields and result_name:
                env = {}
                all_found = True
                for i, f in enumerate(fields):
                    v = metrics.get(f, '')
                    try:
                        v = float(v)
                    except (ValueError, TypeError):
                        all_found = False
                        break
                    env[chr(97 + i)] = v
                if all_found:
                    result = eval(formula, {"__builtins__": {}}, env)
                    if isinstance(result, float) and result == int(result):
                        result = int(result)
                    else:
                        result = round(result, 4)
                    metrics[result_name] = result
                    logger.info(f'  PARSE: 自定义公式 {result_name} = {formula} = {result}')
        except Exception as e:
            logger.info(f'  PARSE: 自定义公式计算失败: {e}')

    # 识别区域名称（跳过位置匹配时不提取）
    skip_location = matched_type.get('skip_location_match', False) if matched_type else False
    ocr_location = ''
    region_name = ''
    if not skip_location:
        ocr_location = _extract_location_from_lines(lines)
        if ocr_location:
            logger.info(f'  PARSE: 从OCR文本提取位置: {ocr_location}')
        region_name = ocr_location
        if not region_name and filename:
            region_name = _extract_location_from_lines([filename])
            if region_name:
                logger.info(f'  PARSE: 从文件名提取位置: {region_name}')

        # 应用位置别名
        if region_name:
            gvars = _load_global_vars()
            aliases = gvars.get('location_aliases', {})
            if region_name in aliases:
                logger.info(f'  PARSE: 位置别名 {region_name} -> {aliases[region_name]}')
                region_name = aliases[region_name]
    else:
        logger.info(f'  PARSE: 跳过位置提取（skip_location_match=true）')


    # 从OCR文本中提取截图日期
    screenshot_date = None
    time_patterns = [
        r'(\d{4})[.\-/](\d{1,2})[.\-/](\d{1,2})',
        r'(\d{4})年(\d{1,2})月(\d{1,2})日',
    ]
    for tp in time_patterns:
        tm = re.search(tp, all_text)
        if tm:
            try:
                y, mo, d = int(tm.group(1)), int(tm.group(2)), int(tm.group(3))
                if 2020 <= y <= 2099 and 1 <= mo <= 12 and 1 <= d <= 31:
                    screenshot_date = f'{y}-{mo:02d}-{d:02d}'
                    logger.info(f'  PARSE: 识别到截图日期 {screenshot_date}')
                    break
            except (ValueError, IndexError):
                pass

    if not metrics:
        return None

    # 构建状态详情（显示 OCR 原始指标和公式计算结果，不显示 _display）
    reverse_label_map = {v: k for k, v in label_map.items()}
    status_detail_parts = []
    for key, val in metrics.items():
        if key.endswith('_display'):
            continue
        label_name = reverse_label_map.get(key, key)
        status_detail_parts.append(f'{label_name}: {val}')
    if online_rate:
        status_detail_parts.append(f'在线率: {online_rate}')

    # 根据 result_rules 判断整体状态
    result = '正常'
    result_rules = matched_type.get('result_rules', {})
    for rule, outcome in result_rules.items():
        # delta 规则：与前一天对比
        m = re.match(r'^(\w+)_?delta([><=]+)([\d.]+)$', rule)
        if m:
            metric_key, op, threshold = m.group(1), m.group(2), float(m.group(3))
            val = metrics.get(metric_key, 0)
            # delta 值会在保存时计算，这里先跳过
            logger.info(f'  PARSE: delta 规则 {rule}（保存时评估）')
            continue
        
        # abs 规则：绝对值阈值
        m = re.match(r'^(\w+)_?abs([><=]+)([\d.]+)$', rule)
        if m:
            metric_key, op, threshold = m.group(1), m.group(2), float(m.group(3))
            val = metrics.get(metric_key, 0)
            if _evaluate_comparison(val, op, threshold):
                result = outcome
                logger.info(f'  PARSE: abs 规则 {rule} -> {outcome}')
            continue

        # 原有规则：直接比较（key 不存在则跳过）
        m = re.match(r'^(\w+)([><=]+)([\d.]+)$', rule)
        if not m:
            continue
        metric_key, op, threshold = m.group(1), m.group(2), float(m.group(3))
        if metric_key not in metrics:
            continue
        val = metrics[metric_key]
        if _evaluate_comparison(val, op, threshold):
            result = outcome
            break

    # 在线率特殊处理
    if online_rate and 'online_rate' in str(result_rules):
        rate_value = float(online_rate.replace('%', ''))
        for rule, outcome in result_rules.items():
            m = re.match(r'^online_rate([><=]+)([\d.]+)$', rule)
            if not m:
                continue
            op, threshold = m.group(1), float(m.group(2))
            if op == '<' and rate_value < threshold:
                result = outcome
                break
            elif op == '<=' and rate_value <= threshold:
                result = outcome
                break

    now = datetime.now().strftime('%Y-%m-%d %H:%M')
    timestamp = f'{screenshot_date} {now.split(" ")[1]}' if screenshot_date else now

    point_name = ''
    if name_from_first_line and unmatched_lines:
        point_name = unmatched_lines[0]
        logger.info(f'  PARSE: 从首行提取名称: {point_name}')
    if not point_name and skip_location:
        point_name = matched_type.get('name', '')
        if point_name:
            logger.info(f'  PARSE: 跳过位置匹配，使用仪表盘名称: {point_name}')
    if not point_name and region_name:
        point_name = region_name
        logger.info(f'  PARSE: point_name 为空，使用位置名称: {point_name}')

    result_data = {
        'point_name': point_name,
        'location': region_name,
        'ocr_location': ocr_location or '',
        'result': result,
        'status_detail': '; '.join(status_detail_parts),
        'notes': '',
        'inspector': '',
        'timestamp': timestamp,
        'is_dashboard': True,
        'dashboard_type': matched_type['id'],
        'dashboard_type_name': matched_type['name'],
        'dashboard_category': matched_type.get('category', matched_type['name']),
        'skip_location_match': matched_type.get('skip_location_match', False),
        'metrics': {k: v for k, v in metrics.items() if not k.endswith('_display')},
        'virtual_keys': _get_virtual_metric_keys_for_type(matched_type),
    }

    return result_data


def parse_traditional_form(lines):
    """解析传统表单格式（兼容原有逻辑）"""
    results = []
    pending_object = ''
    pending_result = ''
    pending_status = ''
    pending_inspector = ''
    pending_time = None

    SKIP_PATTERNS = ['巡检表', '检查表', '签字', '日期', '备注', '说明', '注意事项']

    for line in lines:
        line = line.strip()
        if not line:
            continue

        if any(kw in line for kw in SKIP_PATTERNS) and len(line) < 10:
            continue

        # 尝试识别时间
        time_match = re.search(r'(\d{4}[-/]\d{1,2}[-/]\d{1,2})[\sT]*(\d{1,2}:\d{1,2})?', line)
        if time_match:
            date_str = time_match.group(1)
            time_str = time_match.group(2) or '00:00'
            try:
                pending_time = datetime.strptime(f'{date_str} {time_str}', '%Y-%m-%d %H:%M')
            except ValueError:
                try:
                    pending_time = datetime.strptime(f'{date_str} {time_str}', '%Y/%m/%d %H:%M')
                except ValueError:
                    pass

        # 尝试识别巡检结果
        for result, keywords in RESULT_KEYWORDS.items():
            if any(kw in line.lower() for kw in keywords):
                pending_result = result
                break

        # 尝试识别设备类型
        device_type = None
        for dtype, keywords in DEVICE_TYPES.items():
            if any(kw in line.lower() for kw in keywords):
                device_type = dtype
                break

        # 尝试识别巡检对象名称（包含设备类型关键词或编号）
        has_object_keyword = any(kw in line.lower() for kw in ['服务器', '交换机', '路由器', 'ups', '空调', '监控', '机柜', '设备', 'server', 'switch', 'router'])
        has_number = re.search(r'[A-Za-z]*-\d+|\d+号|\d+号机柜', line)

        if has_object_keyword or has_number:
            pending_object = line

        # 尝试识别巡检人员（包含姓名特征）
        name_match = re.search(r'巡检人[：:]*\s*([^\s]+)', line)
        if name_match:
            pending_inspector = name_match.group(1)
        elif re.match(r'^[\u4e00-\u9fa5]{2,4}$', line) and not pending_inspector:
            # 可能是人名（2-4个汉字）
            pending_inspector = line

        # 状态详情（温度、负载、运行状态等）
        status_keywords = ['温度', '负载', '运行', '状态', 'cpu', '内存', '磁盘', '网络', '功率']
        if any(kw in line.lower() for kw in status_keywords):
            pending_status = line

        # 当收集到完整信息时，保存记录
        if pending_object and pending_result:
            results.append({
                'point_name': pending_object,
                'result': pending_result,
                'status_detail': pending_status,
                'inspector': pending_inspector,
                'timestamp': pending_time or datetime.now(),
            })
            pending_object = ''
            pending_result = ''
            pending_status = ''

    return results


@main.route('/api/export/excel', methods=['GET', 'POST'])
def api_export_excel():
    """导出巡检报告 Excel"""
    data = request.get_json() if request.method == 'POST' else {}
    report_date = data.get('date', request.args.get('date', datetime.now().strftime('%Y%m%d')))
    selected_objects = data.get('objects', None)  # [{name, location, type, label, skipLocation}]
    
    session = SessionLocal()
    dashboard_types = _load_dashboard_types()
    try:
        # 如果有选中的对象，只导出这些
        if selected_objects:
            # 按位置分组（skipLocation的用name，其他的用location）
            groups = {}
            for sel in selected_objects:
                obj_name = sel.get('name', '')
                obj_location = sel.get('location', '')
                skip_loc = sel.get('skipLocation', False)
                filters = {'name': obj_name, 'status': 'active'}
                if obj_location and not skip_loc:
                    filters['location'] = obj_location
                obj = session.query(InspectionObject).filter_by(**filters).first()
                if not obj:
                    continue
                
                group_key = sel.get('label', obj_name)
                if group_key not in groups:
                    groups[group_key] = []
                
                # 获取指定日期的巡检记录
                try:
                    target_date = datetime.strptime(report_date, '%Y%m%d')
                except:
                    target_date = datetime.now()
                
                day_start = target_date.replace(hour=0, minute=0, second=0)
                day_end = target_date.replace(hour=23, minute=59, second=59)
                
                record = (
                    session.query(InspectionRecord)
                    .filter(InspectionRecord.object_id == obj.id,
                            InspectionRecord.timestamp >= day_start,
                            InspectionRecord.timestamp <= day_end)
                    .order_by(InspectionRecord.timestamp.desc())
                    .first()
                )
                
                if not record:
                    record = (
                        session.query(InspectionRecord)
                        .filter_by(object_id=obj.id)
                        .order_by(InspectionRecord.timestamp.desc())
                        .first()
                    )
                
                metrics = {}
                if record and record.metrics:
                    try:
                        metrics = json.loads(record.metrics)
                    except:
                        pass
                
                groups[group_key].append({
                    'object': obj,
                    'record': record,
                    'metrics': metrics,
                    'sort_order': obj.sort_order or 0,
                })
        else:
            # 未选择则导出全部
            objects = session.query(InspectionObject).filter_by(status='active').all()
            groups = {}
            for obj in objects:
                skip_loc = False
                if obj.device_type:
                    for dt in _load_dashboard_types():
                        if dt.get('category') == obj.device_type and dt.get('skip_location_match'):
                            skip_loc = True
                            break
                group_key = obj.name if skip_loc else (obj.location or obj.name)
                if group_key not in groups:
                    groups[group_key] = []
                
                try:
                    target_date = datetime.strptime(report_date, '%Y%m%d')
                except:
                    target_date = datetime.now()
                
                day_start = target_date.replace(hour=0, minute=0, second=0)
                day_end = target_date.replace(hour=23, minute=59, second=59)
                
                record = (
                    session.query(InspectionRecord)
                    .filter(InspectionRecord.object_id == obj.id,
                            InspectionRecord.timestamp >= day_start,
                            InspectionRecord.timestamp <= day_end)
                    .order_by(InspectionRecord.timestamp.desc())
                    .first()
                )
                
                if not record:
                    record = (
                        session.query(InspectionRecord)
                        .filter_by(object_id=obj.id)
                        .order_by(InspectionRecord.timestamp.desc())
                        .first()
                    )
                
                metrics = {}
                if record and record.metrics:
                    try:
                        metrics = json.loads(record.metrics)
                    except:
                        pass
                
                groups[group_key].append({
                    'object': obj,
                    'record': record,
                    'metrics': metrics,
                    'sort_order': obj.sort_order or 0,
                })
        
        # 按 sort_order 排序各组内的对象
        for key in groups:
            groups[key].sort(key=lambda x: x.get('sort_order', x['object'].sort_order or 0))
        
        # 创建 Excel
        wb = Workbook()
        ws = wb.active
        ws.title = '巡检报告'
        
        # 样式定义 - 微软雅黑 14号
        default_font = Font(name='微软雅黑', size=14)
        header_font = Font(name='微软雅黑', size=14, bold=True)
        col_header_font = Font(name='微软雅黑', size=14, bold=True)
        center_align = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
        red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        
        # 固定列：位置/对象, 总数, 在线, 离线, 在线率, 备注
        headers = ['位置/对象', '总数', '在线', '离线', '在线率', '备注']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=header)
            cell.font = col_header_font
            cell.alignment = center_align
            cell.border = thin_border
        
        # 表头
        ws.merge_cells('A1:F1')
        ws['A1'] = '智慧城管平台'
        ws['A1'].font = Font(name='微软雅黑', size=14, bold=True)
        ws['A1'].alignment = center_align
        
        # 数据行
        row = 3
        for group_name, items in groups.items():
            for item in items:
                obj = item['object']
                metrics = item['metrics']
                
                # 获取在线值（从 metrics 中取）
                online_val = 0
                for k in ['在线', 'online']:
                    v = metrics.get(k, '')
                    if v != '':
                        try:
                            online_val = int(float(v))
                        except:
                            online_val = 0
                        break
                
                # 总数：优先使用 metrics 中的 "总数"，无则汇总所有计数类指标
                total = 0
                for k in ['总数', 'total', 'Total']:
                    v = metrics.get(k, '')
                    if v != '':
                        try:
                            total = int(float(v))
                        except:
                            pass
                        break
                if total == 0:
                    # 汇总所有数值型指标（排除百分比/率类和总数类，避免重复计算）
                    skip_keys = {'在线率', 'onlinerate', 'rate', '总数', 'total', 'Total'}
                    for k, v in metrics.items():
                        if k.startswith('_') or k in skip_keys or '总数' in k or 'total' in k.lower():
                            continue
                        try:
                            n = float(str(v).replace('%', '').strip())
                            if n > 0:
                                total += int(n)
                        except:
                            pass
                
                # 离线值 = 总数 - 在线
                offline_val = max(0, total - online_val) if total > 0 else None
                
                # 在线率：优先使用已有的"在线率"指标值
                rate = ''
                rate_val = 0
                rate_from_metrics = None
                for k in ['在线率', 'onlinerate']:
                    v = metrics.get(k, '')
                    if v != '':
                        try:
                            rate_from_metrics = float(v)
                        except:
                            pass
                        break
                if rate_from_metrics is not None:
                    rate_val = rate_from_metrics
                    if 0 < rate_val <= 1:
                        rate_val = rate_val * 100
                    rate = f'{rate_val:.2f}%'
                elif total > 0:
                    rate_val = online_val / total * 100
                    rate = f'{rate_val:.2f}%'
                
                # 从 dashboard_types.json 读取 online_rate 阈值规则
                warn_threshold = 80
                error_threshold = 95
                for dt in dashboard_types:
                    if dt.get('category') == obj.device_type:
                        result_rules = dt.get('result_rules', {})
                        for rule, outcome in result_rules.items():
                            rm = re.match(r'^online_rate([><=]+)([\d.]+)$', rule)
                            if rm:
                                op, threshold = rm.group(1), float(rm.group(2))
                                if op == '<':
                                    error_threshold = threshold
                                elif op == '<=':
                                    error_threshold = threshold
                        break
                
                # 写入数据
                display_name = group_name
                ws.cell(row=row, column=1, value=display_name).font = default_font
                ws.cell(row=row, column=1).border = thin_border
                ws.cell(row=row, column=2, value=total).font = default_font
                ws.cell(row=row, column=2).alignment = center_align
                ws.cell(row=row, column=2).border = thin_border
                ws.cell(row=row, column=3, value=online_val).font = default_font
                ws.cell(row=row, column=3).alignment = center_align
                ws.cell(row=row, column=3).border = thin_border
                ws.cell(row=row, column=4, value=offline_val if offline_val is not None else '').font = default_font
                ws.cell(row=row, column=4).alignment = center_align
                ws.cell(row=row, column=4).border = thin_border
                
                rate_cell = ws.cell(row=row, column=5, value=rate)
                rate_cell.font = default_font
                rate_cell.alignment = center_align
                rate_cell.border = thin_border
                if rate and '%' in str(rate):
                    if rate_val >= error_threshold:
                        rate_cell.fill = green_fill
                    elif rate_val >= warn_threshold:
                        rate_cell.fill = yellow_fill
                    else:
                        rate_cell.fill = red_fill
                
                ws.cell(row=row, column=6, value='').font = default_font
                ws.cell(row=row, column=6).border = thin_border
                row += 1
        
        # 设置列宽
        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 25
        
        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f'{report_date} 巡检报告.xlsx'
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    finally:
        session.close()


@main.route('/')
def index():
    """首页：巡检对象列表"""
    _is_admin = session.get('is_admin')
    _inspector_id = session.get('inspector_id')
    db = SessionLocal()
    try:
        objects = db.query(InspectionObject).filter_by(status='active').order_by(InspectionObject.sort_order, InspectionObject.id).all()
        # 获取每个对象的最近巡检记录和指标
        object_data = []
        stats = {'total': len(objects), 'normal': 0, 'warning': 0, 'error': 0, 'no_record': 0}
        today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
        
        for obj in objects:
            baseq = db.query(InspectionRecord).filter_by(object_id=obj.id)
            if not _is_admin and _inspector_id:
                baseq = baseq.filter(InspectionRecord.inspector_id.in_([_inspector_id, None]))
            latest_record = baseq.order_by(InspectionRecord.timestamp.desc()).first()
            # 获取今天的巡检记录
            today_baseq = db.query(InspectionRecord).filter(
                InspectionRecord.object_id == obj.id,
                InspectionRecord.timestamp >= today)
            if not _is_admin and _inspector_id:
                today_baseq = today_baseq.filter(InspectionRecord.inspector_id.in_([_inspector_id, None]))
            today_record = today_baseq.order_by(InspectionRecord.timestamp.desc()).first()
            # 获取指标配置
            metrics = db.query(ObjectMetric).filter_by(object_id=obj.id).all()
            metric_list = [{'key': m.key, 'name': m.name, 'unit': m.unit, 'show_in_chart': m.show_in_chart} for m in metrics]
            
            # 解析最新记录的指标值
            latest_metrics = {}
            if latest_record and latest_record.metrics:
                try:
                    latest_metrics = json.loads(latest_record.metrics)
                except:
                    pass
            
            # 判断是否跳过位置匹配
            skip_location = False
            if obj.device_type:
                dashboard_types = _load_dashboard_types()
                for dt in dashboard_types:
                    if dt.get('category') == obj.device_type and dt.get('skip_location_match'):
                        skip_location = True
                        break
            
            # 根据当前阈值重新评估最新记录状态（热更新）
            if latest_record:
                parsed = _parse_status_to_metrics(latest_record.status_detail or '')
                if latest_record.metrics:
                    try:
                        parsed.update(json.loads(latest_record.metrics))
                    except:
                        pass
                reevaluated = _check_metrics_thresholds(parsed, obj.id, db)
                if reevaluated and latest_record.result != reevaluated:
                    latest_record.result = reevaluated
            
            # 统计状态（按今日记录统计）
            if not today_record:
                stats['no_record'] += 1
            elif today_record.result == '正常':
                stats['normal'] += 1
            elif today_record.result == '异常':
                stats['error'] += 1
            else:
                stats['warning'] += 1
            
            object_data.append({
                'object': obj,
                'latest_record': latest_record,
                'today_record': today_record,
                'metric_list': metric_list,
                'latest_metrics': latest_metrics,
                'skip_location': skip_location,
                'sort_order': obj.sort_order or 0,
            })
        # 已巡检的排前面，未巡检的排后面；各自分组内按 sort_order 排序
        # 如果全部已巡检，按 异常 > 需关注 > 正常 排序
        has_uninspected = any(not x['today_record'] for x in object_data)
        _status_priority = {'异常': 0, '需关注': 1, '正常': 2}
        if has_uninspected:
            object_data.sort(key=lambda x: (0 if x['today_record'] else 1, x['sort_order']))
        else:
            object_data.sort(key=lambda x: (_status_priority.get(x['today_record'].result, 9), x['sort_order']))
        # 收集所有不重复的 device_type 和 location
        all_types = sorted(set(
            obj.device_type for obj in objects if obj.device_type
        ))
        all_locations = sorted(set(
            obj.location for obj in objects if obj.location
        ))
        return render_template('index.html', object_data=object_data,
                               all_device_types=all_types, all_locations=all_locations,
                               stats=stats)
    finally:
        db.close()


@main.route('/object/<int:object_id>')
def object_detail(object_id):
    """巡检对象详情页"""
    _is_admin = session.get('is_admin')
    _inspector_id = session.get('inspector_id')
    db = SessionLocal()
    try:
        obj = db.query(InspectionObject).filter_by(id=object_id).first()
        if not obj:
            return "巡检对象不存在", 404

        q = db.query(InspectionRecord).filter_by(object_id=object_id)
        if not _is_admin and _inspector_id:
            q = q.filter(InspectionRecord.inspector_id.in_([_inspector_id, None]))
        records = q.order_by(InspectionRecord.timestamp.desc()).all()

        all_inspectors = db.query(Inspector).all()
        default_inspector = all_inspectors[0] if all_inspectors else None

        records_data = [
            {
                'id': r.id,
                'timestamp': r.timestamp.isoformat(),
                'result': r.result,
                'status_detail': r.status_detail,
                'metrics': json.loads(r.metrics) if r.metrics else {},
                'notes': r.notes,
                'inspector': r.inspector.name if r.inspector else (default_inspector.name if default_inspector else '未知'),
            }
            for r in records
        ]

        # 统计数据
        total_records = len(records)
        normal_count = sum(1 for r in records if r.result == '正常')
        abnormal_count = sum(1 for r in records if r.result == '异常')
        warn_count = sum(1 for r in records if r.result == '需关注')

        # 按日去重，每日期取最新一条记录（用于趋势图）
        daily_latest = {}
        for record in records:
            day_key = record.timestamp.strftime('%Y-%m-%d')
            if day_key not in daily_latest:
                daily_latest[day_key] = record

        # 按结果分组，用于图表（跳过周末）
        result_timeline = {}
        for record in daily_latest.values():
            # 跳过周末（周六=5，周日=6）
            if record.timestamp.weekday() >= 5:
                continue
            result_key = record.result
            if result_key not in result_timeline:
                result_timeline[result_key] = []
            ts_ms = int(record.timestamp.timestamp() * 1000)
            status_y = {'正常': 1, '异常': 0, '需关注': 0.5}
            result_timeline[result_key].append({
                'x': ts_ms,
                'y': status_y.get(record.result, 0.5),
                'result': record.result,
                'inspector': record.inspector.name if record.inspector else (default_inspector.name if default_inspector else '未知'),
            })

        # 获取指标配置
        object_metrics = db.query(ObjectMetric).filter_by(object_id=object_id).order_by(ObjectMetric.sort_order).all()
        # 收集增量指标的 result_name
        increment_keys = set()
        for dt in _load_dashboard_types():
            if dt.get('category') == obj.device_type:
                for c in (dt.get('calc_configs') or []):
                    if c.get('type') == 'increment' and c.get('result_name'):
                        increment_keys.add(c['result_name'])
        metrics_config = [
            {
                'id': m.id, 'key': m.key, 'name': m.name, 'unit': m.unit,
                'max_value': m.max_value, 'show_in_chart': m.show_in_chart,
                'warn_threshold': m.warn_threshold, 'error_threshold': m.error_threshold,
                'threshold_direction': m.threshold_direction
            }
            for m in object_metrics
        ]

        return render_template(
            'object_detail.html',
            object=obj,
            records=records,
            records_data=records_data,
            total_records=total_records,
            normal_count=normal_count,
            abnormal_count=abnormal_count,
            warn_count=warn_count,
            result_timeline=result_timeline,
            metrics_config=metrics_config
        )
    finally:
        db.close()


@main.route('/api/inspection_history/<int:object_id>')
def api_inspection_history(object_id):
    """巡检历史 API"""
    _is_admin = session.get('is_admin')
    _inspector_id = session.get('inspector_id')
    db = SessionLocal()
    try:
        q = db.query(InspectionRecord).filter_by(object_id=object_id)
        if not _is_admin and _inspector_id:
            q = q.filter(InspectionRecord.inspector_id.in_([_inspector_id, None]))
        records = q.order_by(InspectionRecord.timestamp.desc()).all()

        all_inspectors = db.query(Inspector).all()

        data = [
            {
                'id': r.id,
                'timestamp': r.timestamp.isoformat(),
                'result': r.result,
                'status_detail': r.status_detail,
                'metrics': json.loads(r.metrics) if r.metrics else {},
                'notes': r.notes,
                'inspector': r.inspector.name if r.inspector else (all_inspectors[0].name if all_inspectors else '未知'),
            }
            for r in records
        ]

        return jsonify(data)
    finally:
        db.close()


@main.route('/api/inspection_history/delete/<int:record_id>', methods=['POST'])
def api_inspection_history_delete(record_id):
    """删除巡检记录"""
    session = SessionLocal()
    try:
        record = session.query(InspectionRecord).get(record_id)
        if not record:
            return jsonify({'error': '记录不存在'}), 404
        object_id = record.object_id
        session.delete(record)
        session.commit()
        return jsonify({'ok': True, 'object_id': object_id})
    except Exception as e:
        session.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        session.close()


@main.route('/api/records/backfill', methods=['POST'])
def api_records_backfill():
    """回填历史记录：补全缺失的"总数"和"离线"字段（仅从已有完整记录推算，不使用 max_value）"""
    session = SessionLocal()
    try:
        objects = session.query(InspectionObject).filter_by(status='active').all()
        updated = 0
        for obj in objects:
            records = (
                session.query(InspectionRecord)
                .filter_by(object_id=obj.id)
                .order_by(InspectionRecord.timestamp.desc())
                .all()
            )
            if not records:
                continue

            # 仅从已有记录中获取已知总数（不使用 max_value）
            known_total = 0
            for r in records:
                m = json.loads(r.metrics) if r.metrics else {}
                for k in ['总数', 'total', 'Total']:
                    v = m.get(k, '')
                    if v != '':
                        try:
                            known_total = int(float(v))
                        except:
                            pass
                        break
                if known_total > 0:
                    break

            for r in records:
                m = json.loads(r.metrics) if r.metrics else {}
                changed = False

                # 补全总数
                has_total = False
                for k in ['总数', 'total', 'Total']:
                    if m.get(k, '') != '':
                        has_total = True
                        break
                if not has_total and known_total > 0:
                    m['总数'] = known_total
                    changed = True

                # 补全离线 = 总数 - 在线
                online_val = None
                for k in ['在线', 'online']:
                    v = m.get(k, '')
                    if v != '':
                        try:
                            online_val = int(float(v))
                        except:
                            pass
                        break
                has_offline = False
                for k in ['离线', 'offline']:
                    if m.get(k, '') != '':
                        has_offline = True
                        break
                total_val = None
                for k in ['总数', 'total', 'Total']:
                    v = m.get(k, '')
                    if v != '':
                        try:
                            total_val = int(float(v))
                        except:
                            pass
                        break
                if not has_offline and online_val is not None and total_val and total_val > 0:
                    m['离线'] = max(0, total_val - online_val)
                    changed = True

                if changed:
                    r.metrics = json.dumps(m, ensure_ascii=False)
                    updated += 1

        session.commit()
        return jsonify({'ok': True, 'updated': updated})
    except Exception as e:
        session.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        session.close()


@main.route('/api/records/backfill-increment', methods=['POST'])
def api_records_backfill_increment():
    """回填增量指标：按时间顺序遍历记录，逐条计算与上一条的差值"""
    session = SessionLocal()
    try:
        types_list = _load_dashboard_types()
        # 收集所有 increment 配置: { source_field, result_name }
        increment_configs = []
        for t in types_list:
            calc_configs = t.get('calc_configs') or []
            single = t.get('calc_config')
            if single and single.get('type') == 'increment':
                calc_configs = [single] + calc_configs
            for c in calc_configs:
                if c.get('type') == 'increment' and c.get('source_field') and c.get('result_name'):
                    increment_configs.append(c)
        if not increment_configs:
            return jsonify({'ok': True, 'updated': 0, 'message': '无增量配置'})

        objects = session.query(InspectionObject).filter_by(status='active').all()
        updated = 0
        for obj in objects:
            records = (
                session.query(InspectionRecord)
                .filter_by(object_id=obj.id)
                .order_by(InspectionRecord.timestamp.asc())
                .all()
            )
            if not records:
                continue
            for i in range(len(records)):
                curr_r = records[i]
                curr_m = json.loads(curr_r.metrics) if curr_r.metrics else {}
                changed = False
                for cfg in increment_configs:
                    src = cfg['source_field']
                    rname = cfg['result_name']
                    curr_val = curr_m.get(src)
                    if curr_val is None:
                        continue
                    if i == 0:
                        # 首条记录，无对比基准，增量为0
                        increment = 0
                    else:
                        prev_r = records[i - 1]
                        prev_m = json.loads(prev_r.metrics) if prev_r.metrics else {}
                        pv = prev_m.get(src)
                        try:
                            curr_num = float(str(curr_val).rstrip('%'))
                            prev_num = float(str(pv).rstrip('%')) if pv is not None else 0
                            increment = curr_num - prev_num
                            # 按天数平滑：间隔>1天时除以天数
                            gap_days = (curr_r.timestamp - prev_r.timestamp).total_seconds() / 86400
                            if gap_days > 1:
                                increment = increment / gap_days
                        except (ValueError, TypeError):
                            continue
                    increment = round(increment, 1)
                    curr_m[rname] = increment
                    changed = True
                if changed:
                    curr_r.metrics = json.dumps(curr_m, ensure_ascii=False)
                    updated += 1
        session.commit()
        return jsonify({'ok': True, 'updated': updated})
    except Exception as e:
        session.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        session.close()


@main.route('/api/records/cleanup', methods=['POST'])
def api_records_cleanup():
    """清理错误的"总数"和"离线"：移除所有记录中的这两个字段，之后可重新回填"""
    session = SessionLocal()
    try:
        records = session.query(InspectionRecord).all()
        cleaned = 0
        for r in records:
            if not r.metrics:
                continue
            m = json.loads(r.metrics)
            changed = False
            for k in ['总数', 'total', 'Total', '离线', 'offline']:
                if k in m:
                    del m[k]
                    changed = True
            if changed:
                r.metrics = json.dumps(m, ensure_ascii=False)
                cleaned += 1
        session.commit()
        return jsonify({'ok': True, 'cleaned': cleaned})
    except Exception as e:
        session.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        session.close()


@main.route('/api/records/compute', methods=['POST'])
def api_records_compute():
    """计算补全：根据仪表盘类型的计算配置批量计算新指标
    请求体 JSON:
    {
        "configs": [
            {"calc_type": "sum", "result_name": "在线", "fields": ["停车", "行驶"]},
            {"calc_type": "percentage", "result_name": "在线率", "numerator": "在线", "denominator": "总数", "denominator_type": "fixed", "denominator_fixed_value": "100"},
            {"calc_type": "difference", "result_name": "离线", "minuend": "总数", "minuend_type": "fixed", "minuend_fixed": "100", "subtrahend": "在线"}
        ]
    }
    """
    data = request.get_json()
    if not data or not data.get('configs'):
        return jsonify({'error': '请提供计算配置'}), 400

    configs = data['configs']
    session = SessionLocal()
    try:
        records = session.query(InspectionRecord).all()
        # 预加载所有对象的 device_type 用于类型匹配
        obj_types = {o.id: o.device_type for o in session.query(InspectionObject).all()}
        total_updated = 0
        results_summary = []

        def _num(val):
            try:
                return float(val)
            except:
                return None

        for cfg in configs:
            calc_type = cfg.get('calc_type', '')
            result_name = cfg.get('result_name', '')
            if not result_name:
                continue
            cfg_category = cfg.get('category', '')

            updated = 0
            for r in records:
                if not r.metrics:
                    continue
                # 跳过类型不匹配的记录
                if cfg_category:
                    obj_type = obj_types.get(r.object_id)
                    if obj_type and obj_type != cfg_category:
                        continue
                m = json.loads(r.metrics)
                result = None

                if calc_type == 'sum':
                    fields = cfg.get('fields', [])
                    vals = []
                    for f in fields:
                        v = _num(m.get(f, ''))
                        if v is None:
                            vals = []
                            break
                        vals.append(v)
                    if vals:
                        result = sum(vals)

                elif calc_type == 'percentage':
                    num_key = cfg.get('numerator', '')
                    den_type = cfg.get('denominator_type', 'field')
                    den_key = cfg.get('denominator', '')
                    num_val = _num(m.get(num_key, ''))
                    if den_type == 'fixed':
                        den_val = _num(cfg.get('denominator_fixed_value', ''))
                    else:
                        den_val = _num(m.get(den_key, ''))
                    decimal_places = cfg.get('decimal_places', 2)
                    if num_val is not None and den_val is not None and den_val != 0:
                        result = round(num_val / den_val * 100, decimal_places)

                elif calc_type == 'difference':
                    minuend_type = cfg.get('minuend_type', 'field')
                    minuend_key = cfg.get('minuend', '')
                    subtrahend_key = cfg.get('subtrahend', '')
                    if minuend_type == 'fixed':
                        left_val = _num(cfg.get('minuend_fixed', ''))
                    else:
                        left_val = _num(m.get(minuend_key, ''))
                    right_val = _num(m.get(subtrahend_key, ''))
                    if left_val is not None and right_val is not None:
                        result = round(left_val - right_val, 2)

                elif calc_type == 'custom':
                    fields = cfg.get('fields', [])
                    formula = cfg.get('formula', '')
                    if formula and fields:
                        env = {}
                        all_found = True
                        for i, f in enumerate(fields):
                            v = _num(m.get(f, ''))
                            if v is None:
                                all_found = False
                                break
                            env[chr(97 + i)] = v
                        if all_found:
                            try:
                                result = round(eval(formula, {"__builtins__": {}}, env), 4)
                            except:
                                result = None

                if result is not None:
                    # 百分比类型自动添加 % 后缀
                    suffix = '%' if calc_type == 'percentage' else ''
                    m[result_name] = f'{result}{suffix}'
                    r.metrics = json.dumps(m, ensure_ascii=False)
                    updated += 1
                    # 确保该对象存在此指标配置（图表用）
                    existing = session.query(ObjectMetric).filter_by(
                        object_id=r.object_id, key=result_name
                    ).first()
                    if not existing:
                        show_in_chart = cfg.get('show_chart', False)
                        max_value = cfg.get('max_value', None)
                        session.add(ObjectMetric(
                            object_id=r.object_id,
                            key=result_name,
                            name=result_name,
                            unit='',
                            max_value=max_value or 100,
                            show_in_chart=show_in_chart
                        ))
                        logger.info(f'  COMPUTE: 自动创建指标配置 key={result_name} show_in_chart={show_in_chart} max_value={max_value}')

            total_updated += updated
            results_summary.append(f'{result_name}: {updated} 条')

        session.commit()
        return jsonify({'ok': True, 'updated': total_updated, 'details': results_summary})
    except Exception as e:
        session.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        session.close()


@main.route('/objects')
def objects():
    """巡检对象管理页面"""
    session = SessionLocal()
    try:
        object_list = session.query(InspectionObject).order_by(InspectionObject.sort_order, InspectionObject.location, InspectionObject.name).all()
        from datetime import timedelta
        for obj in object_list:
            if obj.created_at:
                obj.created_at = obj.created_at + timedelta(hours=8)
        
        return render_template('objects.html', objects=object_list)
    finally:
        session.close()


@main.route('/objects/add', methods=['POST'])
def object_add():
    """添加巡检对象"""
    name = request.form.get('name', '').strip()
    location = request.form.get('location', '').strip() or None
    device_type = request.form.get('device_type', '').strip() or None
    description = request.form.get('description', '').strip() or None
    project_url = request.form.get('project_url', '').strip() or None

    if not name:
        flash('名称不能为空', 'danger')
        return redirect(url_for('main.objects'))

    session = SessionLocal()
    try:
        obj = InspectionObject(name=name, location=location, device_type=device_type, description=description, project_url=project_url)
        session.add(obj)
        session.commit()
        # 双向同步
        if device_type:
            types_list = _load_dashboard_types()
            matched_type = next((t for t in types_list if t.get('category') == device_type or t.get('name') == device_type), None)
            if matched_type:
                _sync_bidirectional(obj.id, matched_type['id'], session)
                session.commit()
        flash(f'巡检对象 "{name}" 添加成功', 'success')
    except Exception as e:
        session.rollback()
        flash(f'添加失败: {e}', 'danger')
    finally:
        session.close()
    return redirect(url_for('main.objects'))


@main.route('/objects/edit/<int:object_id>', methods=['POST'])
def object_edit(object_id):
    """编辑巡检对象"""
    session = SessionLocal()
    try:
        obj = session.query(InspectionObject).get(object_id)
        if not obj:
            flash('巡检对象不存在', 'danger')
            return redirect(url_for('main.objects'))

        name = request.form.get('name', '').strip()
        location = request.form.get('location', '').strip() or None
        device_type = request.form.get('device_type', '').strip() or None
        status = request.form.get('status', '').strip() or 'active'
        description = request.form.get('description', '').strip() or None
        project_url = request.form.get('project_url', '').strip() or None

        if name:
            obj.name = name
        obj.location = location
        obj.device_type = device_type
        obj.status = status
        obj.description = description
        obj.project_url = project_url

        session.commit()
        # 双向同步
        if device_type:
            types_list = _load_dashboard_types()
            matched_type = next((t for t in types_list if t.get('category') == device_type or t.get('name') == device_type), None)
            if matched_type:
                _sync_bidirectional(obj.id, matched_type['id'], session)
                session.commit()
        flash(f'巡检对象 "{obj.name}" 已更新', 'success')
    except Exception as e:
        session.rollback()
        flash(f'更新失败: {e}', 'danger')
    finally:
        session.close()
    return redirect(url_for('main.objects'))


@main.route('/objects/delete/<int:object_id>', methods=['POST'])
def object_delete(object_id):
    """删除巡检对象"""
    session = SessionLocal()
    try:
        obj = session.query(InspectionObject).get(object_id)
        if not obj:
            flash('巡检对象不存在', 'danger')
            return redirect(url_for('main.objects'))

        name = obj.name
        # 删除关联的巡检记录
        session.query(InspectionRecord).filter_by(object_id=obj.id).delete()
        session.delete(obj)
        session.commit()
        flash(f'巡检对象 "{name}" 已删除', 'success')
    except Exception as e:
        session.rollback()
        flash(f'删除失败: {e}', 'danger')
    finally:
        session.close()
    return redirect(url_for('main.objects'))


@main.route('/objects/clone/<int:object_id>', methods=['POST'])
def object_clone(object_id):
    """复制巡检对象"""
    session = SessionLocal()
    try:
        src = session.query(InspectionObject).get(object_id)
        if not src:
            flash('巡检对象不存在', 'danger')
            return redirect(url_for('main.objects'))

        new_obj = InspectionObject(
            name=src.name,
            location='',
            device_type=src.device_type,
            description=src.description,
            status=src.status,
            project_url=src.project_url
        )
        session.add(new_obj)
        session.flush()

        # 复制指标配置
        src_metrics = session.query(ObjectMetric).filter_by(object_id=src.id).all()
        for m in src_metrics:
            new_m = ObjectMetric(
                object_id=new_obj.id,
                key=m.key,
                name=m.name,
                unit=m.unit,
                max_value=m.max_value,
                sort_order=m.sort_order,
                show_in_chart=m.show_in_chart,
                warn_threshold=m.warn_threshold,
                error_threshold=m.error_threshold,
                threshold_direction=m.threshold_direction
            )
            session.add(new_m)

        session.commit()
        flash(f'已复制 "{src.name}"，请修改位置', 'success')
    except Exception as e:
        session.rollback()
        flash(f'复制失败: {e}', 'danger')
    finally:
        session.close()
    return redirect(url_for('main.objects'))


@main.route('/api/objects/list')
def api_objects_list():
    """巡检对象列表 API"""
    session = SessionLocal()
    try:
        objects = session.query(InspectionObject).filter_by(status='active').all()
        return jsonify([
            {
                'id': o.id,
                'name': o.name,
                'location': o.location,
                'device_type': o.device_type,
                'description': o.description
            }
            for o in objects
        ])
    finally:
        session.close()


@main.route('/api/objects/suggestions')
def api_objects_suggestions():
    """获取巡检对象建议（用于自动补全）"""
    query = request.args.get('q', '').strip().lower()
    field = request.args.get('field', 'name')  # name, location, device_type
    
    if not query or len(query) < 1:
        return jsonify([])
    
    session = SessionLocal()
    try:
        suggestions = set()
        
        if field == 'name':
            # 搜索名称
            for row in session.query(InspectionObject.name).filter(
                InspectionObject.status == 'active',
                InspectionObject.name.ilike(f'%{query}%')
            ).distinct().limit(10):
                if row[0]:
                    suggestions.add(row[0])
        
        elif field == 'location':
            # 搜索位置
            for row in session.query(InspectionObject.location).filter(
                InspectionObject.status == 'active',
                InspectionObject.location.ilike(f'%{query}%'),
                InspectionObject.location.isnot(None),
                InspectionObject.location != ''
            ).distinct().limit(10):
                if row[0]:
                    suggestions.add(row[0])
        
        elif field == 'device_type':
            # 搜索设备类型
            # 先从数据库获取
            for row in session.query(InspectionObject.device_type).filter(
                InspectionObject.status == 'active',
                InspectionObject.device_type.ilike(f'%{query}%'),
                InspectionObject.device_type.isnot(None),
                InspectionObject.device_type != ''
            ).distinct().limit(10):
                if row[0]:
                    suggestions.add(row[0])
            
            # 再从仪表盘类型获取（包括 category 和 name）
            types_list = _load_dashboard_types()
            for t in types_list:
                cat = t.get('category', '')
                if cat and query in cat.lower():
                    suggestions.add(cat)
                name = t.get('name', '')
                if name and query in name.lower():
                    suggestions.add(name)
        
        return jsonify(sorted(suggestions))
    finally:
        session.close()


@main.route('/api/objects/sort', methods=['POST'])
def api_objects_sort():
    """保存巡检对象排序"""
    data = request.get_json()
    if not data or 'order' not in data:
        return jsonify({'error': '请提供排序数据'}), 400
    
    session = SessionLocal()
    try:
        order = data['order']  # [{id: 1}, {id: 2}, ...]
        for idx, item in enumerate(order):
            obj_id = item.get('id')
            if obj_id:
                obj = session.query(InspectionObject).get(obj_id)
                if obj:
                    obj.sort_order = idx
        session.commit()
        return jsonify({'ok': True})
    except Exception as e:
        session.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        session.close()


@main.route('/api/points/list')
def api_points_list():
    """巡检点列表 API（兼容前端旧接口，实际返回巡检对象）"""
    session = SessionLocal()
    try:
        objects = session.query(InspectionObject).filter_by(status='active').all()
        return jsonify([
            {
                'id': o.id,
                'name': o.name,
                'location': o.location,
                'device_type': o.device_type,
                'description': o.description
            }
            for o in objects
        ])
    finally:
        session.close()


@main.route('/api/objects/<int:object_id>/metrics', methods=['GET'])
def api_object_metrics_list(object_id):
    """获取巡检对象的指标配置"""
    session = SessionLocal()
    try:
        obj = session.query(InspectionObject).get(object_id)
        metrics = session.query(ObjectMetric).filter_by(object_id=object_id).order_by(ObjectMetric.sort_order).all()
        virtual_keys = set()
        if obj:
            for dt in _load_dashboard_types():
                if dt.get('category') == obj.device_type:
                    virtual_keys.update(_get_virtual_metric_keys_for_type(dt))
        return jsonify([
            {
                'id': m.id,
                'key': m.key,
                'name': m.name,
                'unit': m.unit,
                'max_value': m.max_value,
                'show_in_chart': m.show_in_chart,
                'sort_order': m.sort_order,
                'warn_threshold': m.warn_threshold,
                'error_threshold': m.error_threshold,
                'threshold_direction': m.threshold_direction,
                'is_virtual': m.key in virtual_keys,
            }
            for m in metrics
        ])
    finally:
        session.close()


@main.route('/api/objects/<int:object_id>/metrics', methods=['POST'])
def api_object_metrics_add(object_id):
    """添加指标配置"""
    data = request.get_json()
    if not data:
        return jsonify({'error': '无效数据'}), 400
    key = data.get('key', '').strip()
    name = data.get('name', '').strip()
    unit = data.get('unit', '').strip()
    show_in_chart = data.get('show_in_chart', True)
    sort_order = data.get('sort_order', 0)
    max_value = data.get('max_value', 100)
    warn_threshold = data.get('warn_threshold')
    error_threshold = data.get('error_threshold')
    threshold_direction = data.get('threshold_direction', 'lt')
    if not key or not name:
        return jsonify({'error': 'key 和 name 不能为空'}), 400
    session = SessionLocal()
    try:
        obj = session.query(InspectionObject).get(object_id)
        # 检查 key 是否已存在
        existing = session.query(ObjectMetric).filter_by(object_id=object_id, key=key).first()
        if existing:
            session.close()
            return jsonify({'error': f'key "{key}" 已存在，请使用其他名称'}), 409
        # 不允许手动添加虚拟指标 key
        virtual_keys = _get_virtual_metric_keys_for_object(obj) if obj else set()
        if key in virtual_keys:
            session.close()
            return jsonify({'error': f'key "{key}" 为计算类指标，由仪表盘类型自动管理'}), 400
        metric = ObjectMetric(
            object_id=object_id, key=key, name=name,
            unit=unit, max_value=max_value, show_in_chart=show_in_chart, sort_order=sort_order,
            warn_threshold=warn_threshold, error_threshold=error_threshold, threshold_direction=threshold_direction
        )
        session.add(metric)
        session.commit()
        return jsonify({'id': metric.id, 'key': metric.key, 'name': metric.name,
                        'unit': metric.unit, 'max_value': metric.max_value, 'show_in_chart': metric.show_in_chart,
                        'warn_threshold': metric.warn_threshold, 'error_threshold': metric.error_threshold,
                        'threshold_direction': metric.threshold_direction})
    finally:
        session.close()


@main.route('/api/objects/<int:object_id>/metrics/<int:metric_id>', methods=['PUT'])
def api_object_metrics_update(object_id, metric_id):
    """更新指标配置"""
    data = request.get_json()
    session = SessionLocal()
    try:
        metric = session.query(ObjectMetric).filter_by(id=metric_id, object_id=object_id).first()
        if not metric:
            return jsonify({'error': '指标不存在'}), 404
        # 虚拟指标保护（来自 calc_config 的计算结果）
        obj = session.query(InspectionObject).get(object_id)
        virtual_keys = _get_virtual_metric_keys_for_object(obj) if obj else set()
        is_virtual = metric.key in virtual_keys
        if is_virtual:
            # 只允许编辑阈值和排序，不允许编辑基础配置
            allowed = {'sort_order', 'warn_threshold', 'error_threshold', 'threshold_direction'}
            forbidden = set(data.keys()) - allowed
            if forbidden:
                return jsonify({'error': f'计算类指标仅可编辑阈值和排序，不允许修改: {", ".join(forbidden)}'}), 400
        for field in ('key', 'name', 'unit'):
            if field in data and not is_virtual:
                setattr(metric, field, data[field])
        if 'show_in_chart' in data and not is_virtual:
            metric.show_in_chart = data['show_in_chart']
        if 'max_value' in data and not is_virtual:
            metric.max_value = data['max_value']
        if 'sort_order' in data:
            metric.sort_order = data['sort_order']
        if 'warn_threshold' in data:
            metric.warn_threshold = data['warn_threshold']
        if 'error_threshold' in data:
            metric.error_threshold = data['error_threshold']
        if 'threshold_direction' in data:
            metric.threshold_direction = data['threshold_direction']
        session.commit()
        return jsonify({'ok': True})
    finally:
        session.close()


@main.route('/api/objects/<int:object_id>/metrics/sort', methods=['POST'])
def api_object_metrics_sort(object_id):
    """保存指标排序"""
    data = request.get_json()
    order = data.get('order', [])
    session = SessionLocal()
    try:
        for idx, item in enumerate(order):
            metric_id = item.get('id')
            if metric_id:
                metric = session.query(ObjectMetric).filter_by(id=metric_id, object_id=object_id).first()
                if metric:
                    metric.sort_order = idx
        session.commit()
        return jsonify({'ok': True})
    finally:
        session.close()


@main.route('/api/objects/<int:object_id>/metrics/<int:metric_id>', methods=['DELETE'])
def api_object_metrics_delete(object_id, metric_id):
    """删除指标配置并同步到仪表盘类型"""
    session = SessionLocal()
    try:
        metric = session.query(ObjectMetric).filter_by(id=metric_id, object_id=object_id).first()
        if not metric:
            return jsonify({'error': '指标不存在'}), 404

        # 保护虚拟指标（来自 calc_config 的计算结果）
        obj = session.query(InspectionObject).get(object_id)
        virtual_keys = _get_virtual_metric_keys_for_object(obj) if obj else set()
        if metric.key in virtual_keys:
            return jsonify({'error': '计算类指标不可删除，如需移除请在仪表盘类型管理页面移除计算配置'}), 400

        metric_name = metric.name
        metric_key = metric.key
        session.delete(metric)
        session.commit()

        # 同步到仪表盘类型：删除对应的标签
        obj = session.query(InspectionObject).get(object_id)
        if obj:
            types_list = _load_dashboard_types()
            matched_type = next((t for t in types_list if t['name'] == obj.name or t.get('category') == obj.device_type), None)
            if matched_type and metric_name in matched_type.get('labels', {}):
                del matched_type['labels'][metric_name]
                _save_dashboard_types(types_list)
                logger.info(f'  SYNC: 从仪表盘类型删除标签 {metric_name}({metric_key})')

        return jsonify({'ok': True})
    finally:
        session.close()


@main.route('/api/objects/<int:object_id>/re-evaluate', methods=['POST'])
def api_re_evaluate_records(object_id):
    """根据当前阈值配置重新评估该对象所有巡检记录的异常状态"""
    session = SessionLocal()
    try:
        records = session.query(InspectionRecord).filter_by(object_id=object_id).all()
        if not records:
            return jsonify({'updated': 0})
        updated = 0
        for record in records:
            parsed_metrics = _parse_status_to_metrics(record.status_detail or '')
            if record.metrics:
                try:
                    stored = json.loads(record.metrics) if isinstance(record.metrics, str) else record.metrics
                    parsed_metrics.update(stored)
                except Exception:
                    pass
            threshold_result = _check_metrics_thresholds(parsed_metrics, object_id, session)
            if threshold_result is not None and record.result != threshold_result:
                record.result = threshold_result
                updated += 1
        session.commit()
        return jsonify({'updated': updated})
    except Exception as e:
        session.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        session.close()


@main.route('/api/reevaluate', methods=['POST'])
def api_reevaluate():
    """根据对象阈值重新评估状态"""
    data = request.get_json()
    if not data or not data.get('object_id') or not data.get('metrics'):
        return jsonify({'error': '缺少参数'}), 400
    session = SessionLocal()
    try:
        result = _check_metrics_thresholds(data['metrics'], data['object_id'], session)
        return jsonify({'ok': True, 'result': result or '正常'})
    finally:
        session.close()


@main.route('/daily-list/<int:object_id>')
def daily_list(object_id):
    """日列表历史页面（群成员、粉丝列表等）"""
    session = SessionLocal()
    try:
        obj = session.query(InspectionObject).get(object_id)
        if not obj:
            flash('对象不存在', 'error')
            return redirect(url_for('main.index'))
        
        # 获取日列表记录，按日期倒序
        records = session.query(DailyListRecord).filter_by(
            object_id=object_id
        ).order_by(DailyListRecord.date.desc()).limit(90).all()
        
        # 解析每条记录的 items
        record_data = []
        for r in records:
            try:
                items = json.loads(r.items) if r.items else []
            except:
                items = []
            record_data.append({
                'date': r.date.isoformat() if r.date else '',
                'count': r.count,
                'raw_count': r.raw_count,
                'items': items,
                'created_at': r.created_at.strftime('%H:%M') if r.created_at else ''
            })
        
        return render_template('daily_list.html', obj=obj, records=record_data)
    finally:
        session.close()


@main.route('/api/daily-list/<int:object_id>')
def api_daily_list(object_id):
    """日列表历史 API"""
    session = SessionLocal()
    try:
        days = request.args.get('days', 30, type=int)
        cutoff_date = date.today() - timedelta(days=days)
        
        records = session.query(DailyListRecord).filter(
            DailyListRecord.object_id == object_id,
            DailyListRecord.date >= cutoff_date
        ).order_by(DailyListRecord.date.desc()).all()
        
        result = []
        for r in records:
            try:
                items = json.loads(r.items) if r.items else []
            except:
                items = []
            result.append({
                'date': r.date.isoformat() if r.date else '',
                'count': r.count,
                'raw_count': r.raw_count,
                'items': items,
            })
        
        return jsonify(result)
    finally:
        session.close()


@main.route('/api/import-group-members', methods=['POST'])
def api_import_group_members():
    """导入群成员文本到日列表"""
    data = request.get_json()
    if not data or 'object_id' not in data or 'items' not in data:
        return jsonify({'error': '缺少参数'}), 400
    
    object_id = data['object_id']
    items = data['items']
    
    if not items:
        return jsonify({'error': '成员列表为空'}), 400
    
    session = SessionLocal()
    try:
        obj = session.query(InspectionObject).get(object_id)
        if not obj:
            return jsonify({'error': '巡检对象不存在'}), 400
        
        today = date.today()
        unique_items = list(dict.fromkeys(items))  # 去重保持顺序
        
        # UPSERT
        existing = session.query(DailyListRecord).filter_by(
            object_id=object_id, date=today, content_type='list'
        ).first()
        
        if existing:
            existing.items = json.dumps(unique_items, ensure_ascii=False)
            existing.count = len(unique_items)
            existing.raw_count = len(items)
        else:
            record = DailyListRecord(
                object_id=object_id,
                date=today,
                content_type='list',
                items=json.dumps(unique_items, ensure_ascii=False),
                count=len(unique_items),
                raw_count=len(items)
            )
            session.add(record)
        
        session.commit()
        logger.info(f'  IMPORT: 群成员导入 obj_id={object_id} count={len(unique_items)}')
        return jsonify({'count': len(unique_items), 'object_id': object_id})
    except Exception as e:
        session.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        session.close()


@main.route('/inspectors')
def inspectors():
    """巡检人员管理页面"""
    session = SessionLocal()
    try:
        inspector_list = session.query(Inspector).order_by(Inspector.team, Inspector.name).all()
        return render_template('inspectors.html', inspectors=inspector_list)
    finally:
        session.close()


@main.route('/inspectors/add', methods=['POST'])
def inspector_add():
    """添加巡检人员"""
    name = request.form.get('name', '').strip()
    team = request.form.get('team', '').strip() or None
    contact = request.form.get('contact', '').strip() or None
    username = request.form.get('username', '').strip() or None
    password = request.form.get('password', '').strip() or None
    is_admin = request.form.get('is_admin') == 'on'

    if not name:
        flash('姓名不能为空', 'danger')
        return redirect(url_for('main.inspectors'))

    session_local = SessionLocal()
    try:
        if username:
            exists = session_local.query(Inspector).filter(Inspector.username == username).first()
            if exists:
                flash(f'登录名 "{username}" 已被使用', 'danger')
                return redirect(url_for('main.inspectors'))
        inspector = Inspector(
            name=name, team=team, contact=contact,
            username=username,
            password=generate_password_hash(password) if password else None,
            is_admin=is_admin
        )
        session_local.add(inspector)
        session_local.commit()
        flash(f'巡检人员 "{name}" 添加成功', 'success')
    except Exception as e:
        session_local.rollback()
        flash(f'添加失败: {e}', 'danger')
    finally:
        session_local.close()
    return redirect(url_for('main.inspectors'))


@main.route('/inspectors/edit/<int:inspector_id>', methods=['POST'])
def inspector_edit(inspector_id):
    """编辑巡检人员"""
    session_local = SessionLocal()
    try:
        inspector = session_local.query(Inspector).get(inspector_id)
        if not inspector:
            flash('巡检人员不存在', 'danger')
            return redirect(url_for('main.inspectors'))

        name = request.form.get('name', '').strip()
        team = request.form.get('team', '').strip() or None
        contact = request.form.get('contact', '').strip() or None
        username = request.form.get('username', '').strip() or None
        password = request.form.get('password', '').strip() or None
        is_admin = request.form.get('is_admin') == 'on'

        if name:
            inspector.name = name
        inspector.team = team
        inspector.contact = contact
        inspector.is_admin = is_admin

        if username is not None:
            exists = session_local.query(Inspector).filter(
                Inspector.username == username, Inspector.id != inspector_id
            ).first()
            if exists:
                flash(f'登录名 "{username}" 已被使用', 'danger')
                return redirect(url_for('main.inspectors'))
            inspector.username = username or None

        if password:
            inspector.password = generate_password_hash(password)

        session_local.commit()
        flash(f'巡检人员 "{inspector.name}" 已更新', 'success')
    except Exception as e:
        session_local.rollback()
        flash(f'更新失败: {e}', 'danger')
    finally:
        session_local.close()
    return redirect(url_for('main.inspectors'))


@main.route('/inspectors/delete/<int:inspector_id>', methods=['POST'])
def inspector_delete(inspector_id):
    """删除巡检人员"""
    session = SessionLocal()
    try:
        inspector = session.query(Inspector).get(inspector_id)
        if not inspector:
            flash('巡检人员不存在', 'danger')
            return redirect(url_for('main.inspectors'))

        name = inspector.name
        session.delete(inspector)
        session.commit()
        flash(f'巡检人员 "{name}" 已删除', 'success')
    except Exception as e:
        session.rollback()
        flash(f'删除失败: {e}', 'danger')
    finally:
        session.close()
    return redirect(url_for('main.inspectors'))


@main.route('/api/inspectors/list')
def api_inspectors_list():
    """巡检人员列表 API"""
    session = SessionLocal()
    try:
        inspectors = session.query(Inspector).all()
        return jsonify([
            {'id': i.id, 'name': i.name, 'team': i.team, 'contact': i.contact}
            for i in inspectors
        ])
    finally:
        session.close()


@main.route('/upload')
def upload():
    """OCR 截图识别页面"""
    return render_template('upload.html')


@main.route('/import')
def bulk_import():
    """批量导入页面"""
    return render_template('bulk_import.html')


@main.route('/api/objects/import', methods=['POST'])
def api_objects_import():
    """批量导入巡检对象"""
    data = request.get_json()
    if not data or 'objects' not in data:
        return jsonify({'error': '没有要导入的数据'}), 400

    objects = data['objects']
    session = SessionLocal()
    imported = 0
    try:
        for item in objects:
            name = item.get('name', '').strip()
            location = item.get('location', '').strip() or None
            device_type = item.get('device_type', '').strip() or None
            description = item.get('description', '').strip() or None

            if not name:
                continue

            # 检查是否已存在同名对象
            existing = session.query(InspectionObject).filter_by(name=name).first()
            if existing:
                continue

            obj = InspectionObject(name=name, location=location, device_type=device_type, description=description)
            session.add(obj)
            imported += 1
        session.commit()
    except Exception as e:
        session.rollback()
        return jsonify({'error': f'导入失败: {str(e)}'}), 500
    finally:
        session.close()

    return jsonify({'imported': imported})


@main.route('/api/inspectors/import', methods=['POST'])
def api_inspectors_import():
    """批量导入人员"""
    data = request.get_json()
    if not data or 'inspectors' not in data:
        return jsonify({'error': '没有要导入的数据'}), 400

    inspectors = data['inspectors']
    session = SessionLocal()
    imported = 0
    try:
        for item in inspectors:
            name = item.get('name', '').strip()
            team = item.get('team', '').strip() or None
            contact = item.get('contact', '').strip() or None

            if not name:
                continue

            # 检查是否已存在同名人员
            existing = session.query(Inspector).filter_by(name=name).first()
            if existing:
                continue

            inspector = Inspector(name=name, team=team, contact=contact)
            session.add(inspector)
            imported += 1
        session.commit()
    except Exception as e:
        session.rollback()
        return jsonify({'error': f'导入失败: {str(e)}'}), 500
    finally:
        session.close()

    return jsonify({'imported': imported})


@main.route('/api/records/import', methods=['POST'])
def api_records_import():
    """批量导入巡检记录"""
    data = request.get_json()
    if not data or 'records' not in data:
        return jsonify({'error': '没有要导入的数据'}), 400

    records = data['records']
    session = SessionLocal()
    imported = 0
    skipped = 0
    created = 0
    try:
        all_objects = session.query(InspectionObject).filter_by(status='active').all()
        all_inspectors = session.query(Inspector).all()

        for item in records:
            object_name = item.get('point_name', '').strip() or item.get('object_name', '').strip()
            timestamp_str = item.get('timestamp', '').strip()
            result = item.get('result', '正常').strip()
            inspector_name = item.get('inspector_name', '').strip()
            status_detail = item.get('status_detail', '').strip() or None
            notes = item.get('notes', '').strip() or None

            if not object_name:
                skipped += 1
                continue

            # 匹配巡检对象
            obj = _match_object(object_name, all_objects)
            if not obj:
                # 自动创建新巡检对象
                obj = InspectionObject(
                    name=object_name,
                    location=object_name,
                    device_type='监控',
                    description=f'批量导入 - {status_detail or ""}',
                    status='active'
                )
                session.add(obj)
                session.flush()
                all_objects.append(obj)
                created += 1

            # 匹配巡检人员
            inspector = None
            if inspector_name:
                inspector = _match_inspector(inspector_name, all_inspectors)
            if not inspector and all_inspectors:
                inspector = all_inspectors[0]

            # 解析时间
            timestamp = None
            if timestamp_str:
                try:
                    timestamp = datetime.strptime(timestamp_str, '%Y-%m-%d %H:%M')
                except ValueError:
                    try:
                        timestamp = datetime.strptime(timestamp_str, '%Y-%m-%d')
                    except ValueError:
                        timestamp = datetime.now()
            else:
                timestamp = datetime.now()

            # 根据指标阈值重新判断结果
            parsed_metrics = _parse_status_to_metrics(status_detail or '')
            threshold_result = _check_metrics_thresholds(parsed_metrics, obj.id, session)
            if threshold_result:
                result = threshold_result

            record = InspectionRecord(
                point_id=obj.id,
                object_id=obj.id,
                inspector_id=inspector.id if inspector else None,
                result=result,
                status_detail=status_detail,
                metrics=json.dumps(_parse_status_to_metrics(status_detail), ensure_ascii=False),
                notes=notes,
                timestamp=timestamp
            )
            session.add(record)
            imported += 1
        session.commit()
    except Exception as e:
        session.rollback()
        return jsonify({'error': f'导入失败: {str(e)}'}), 500
    finally:
        session.close()

    return jsonify({'imported': imported, 'skipped': skipped, 'created': created})


@main.route('/api/ocr', methods=['POST'])
def api_ocr():
    """OCR 识别 API"""
    if not ocr_engine:
        return jsonify({'error': 'OCR 引擎未安装，请运行: pip install rapidocr-onnxruntime'}), 500

    data = request.get_json()
    if not data or 'image' not in data:
        return jsonify({'error': '请提供图片数据'}), 400

    image_data = data['image']
    if ',' in image_data:
        image_data = image_data.split(',', 1)[1]

    filename = data.get('filename', '')  # 获取文件名

    # 提取图片创建时间（EXIF 或文件修改时间）
    last_modified = data.get('last_modified')
    try:
        image_bytes_for_exif = base64.b64decode(image_data)
        file_obj = io.BytesIO(image_bytes_for_exif)
        creation_time = get_image_creation_time(file_obj)
    except Exception:
        creation_time = None
    # 如果EXIF无时间，使用文件修改时间
    if not creation_time and last_modified:
        try:
            # last_modified 是毫秒时间戳
            creation_time = datetime.fromtimestamp(last_modified / 1000)
            logger.info(f'  IMAGE: 使用文件修改时间 {creation_time}')
        except Exception:
            pass
    if not creation_time:
        creation_time = datetime.now()

    config = _load_ocr_config()
    ignore_top = config.get('ignore_top', 0)
    ignore_bottom = config.get('ignore_bottom', 0)

    try:
        image_bytes = base64.b64decode(image_data)
        img = Image.open(io.BytesIO(image_bytes))
        img_height = img.height
        img_width = img.width
        logger.info(f'  IMAGE: 原始图片尺寸 {img_width}x{img_height}, 格式 {img.format}, 模式 {img.mode}')
        top_cutoff = int(img_height * ignore_top / 100) if ignore_top > 0 else 0
        bottom_cutoff = img_height - int(img_height * ignore_bottom / 100) if ignore_bottom > 0 else img_height

        with tempfile.NamedTemporaryFile(suffix='.png', delete=False) as tmp:
            img.save(tmp.name)
            tmp_path = tmp.name
        import time as _time
        t0 = _time.time()
        result, _ = ocr_engine(tmp_path)
        elapsed = _time.time() - t0
        logger.info(f'  IMAGE: OCR耗时 {elapsed:.2f}s, 结果条数 {len(result) if result else 0}')
        Path(tmp_path).unlink(missing_ok=True)
    except Exception as e:
        return jsonify({'error': f'图片处理失败: {str(e)}'}), 400

    filtered_result = []
    for item in (result or []):
        if item and len(item) >= 2:
            bbox = item[0]
            if bbox and len(bbox) >= 4:
                y_center = (bbox[0][1] + bbox[2][1]) / 2
                if y_center < top_cutoff or y_center > bottom_cutoff:
                    continue
            filtered_result.append(item)

    items = parse_inspection_form(filtered_result, filename=filename)

    # 将图片创建时间注入每条记录
    for item in items:
        if item.get('is_dashboard'):
            # 优先使用OCR文本中的日期，其次用图片创建时间
            if not item.get('timestamp') or '截图' in item.get('timestamp', ''):
                item['file_creation_time'] = creation_time.strftime('%Y-%m-%d %H:%M')
            else:
                item['file_creation_time'] = creation_time.strftime('%Y-%m-%d %H:%M')

    return jsonify({
        'items': items,
        'raw_lines': [item[1] for item in filtered_result if item],
    })


@main.route('/ocr-admin')
def ocr_admin():
    """OCR 管理页面（隐藏）"""
    config = _load_ocr_config()
    engine_status = 'running' if ocr_engine else 'not_installed'
    return render_template('ocr_admin.html', config=config, engine_status=engine_status)


@main.route('/api/ocr-config', methods=['GET'])
def api_ocr_config_get():
    """获取当前 OCR 配置"""
    return jsonify(_load_ocr_config())


@main.route('/api/ocr-config', methods=['POST'])
def api_ocr_config_apply():
    """实时应用 OCR 配置（不持久化）"""
    global ocr_engine
    data = request.get_json()
    if not data:
        return jsonify({'error': '请提供配置数据'}), 400

    config = _load_ocr_config()
    config.update(data)

    new_engine = _build_ocr_engine(config)
    if new_engine is None:
        return jsonify({'error': 'OCR 引擎未安装，请运行: pip install rapidocr-onnxruntime'}), 500

    ocr_engine = new_engine
    return jsonify({'ok': True, 'config': config})


@main.route('/api/ocr-config/save', methods=['POST'])
def api_ocr_config_save():
    """保存 OCR 配置到文件"""
    global ocr_engine
    data = request.get_json()
    if not data:
        return jsonify({'error': '请提供配置数据'}), 400

    config = _load_ocr_config()
    config.update(data)

    try:
        with open(OCR_CONFIG_PATH, 'w', encoding='utf-8') as f:
            json.dump(config, f, ensure_ascii=False, indent=2)
    except Exception as e:
        return jsonify({'error': f'保存失败: {str(e)}'}), 500

    new_engine = _build_ocr_engine(config)
    if new_engine is not None:
        ocr_engine = new_engine

    return jsonify({'ok': True, 'config': config})


@main.route('/api/global-vars', methods=['GET'])
def api_global_vars_get():
    """获取全局变量配置"""
    return jsonify(_load_global_vars())


@main.route('/api/global-vars', methods=['POST'])
def api_global_vars_save():
    """保存全局变量配置"""
    data = request.get_json()
    if not data:
        return jsonify({'error': '请提供配置数据'}), 400

    current = _load_global_vars()
    current.update(data)
    try:
        _save_global_vars(current)
    except Exception as e:
        return jsonify({'error': f'保存失败: {str(e)}'}), 500
    return jsonify({'ok': True, 'config': current})


@main.route('/api/dashboard-types', methods=['GET'])
def api_dashboard_types_list():
    """获取所有仪表盘类型"""
    return jsonify(_load_dashboard_types())


@main.route('/api/dashboard-types/calc-configs', methods=['GET'])
def api_dashboard_types_calc_configs():
    """获取所有仪表盘类型的计算配置（用于数据维护计算补全）"""
    types_list = _load_dashboard_types()
    configs = []
    for t in types_list:
        calc = t.get('calc_config')
        if not calc or not calc.get('type') or not calc.get('result_name'):
            continue
        calc_type = calc.get('type')
        cfg = {
            'type_id': t.get('id'),
            'type_name': t.get('name'),
            'category': t.get('category', ''),
            'calc_type': calc_type,
            'result_name': calc.get('result_name'),
            'show_chart': calc.get('show_chart', False),
            'max_value': calc.get('max_value')
        }
        if calc_type == 'sum':
            cfg['fields'] = calc.get('fields', [])
        elif calc_type == 'percentage':
            cfg['numerator'] = calc.get('numerator', '')
            cfg['denominator'] = calc.get('denominator', '')
            cfg['denominator_type'] = calc.get('denominator_type', 'field')
            cfg['decimal_places'] = calc.get('decimal_places', 2)
            cfg['format'] = calc.get('format', 'decimal')
        elif calc_type == 'difference':
            cfg['minuend'] = calc.get('minuend', '')
            cfg['minuend_type'] = calc.get('minuend_type', 'field')
            cfg['minuend_fixed'] = calc.get('minuend_fixed', '')
            cfg['subtrahend'] = calc.get('subtrahend', '')
        elif calc_type == 'custom':
            cfg['fields'] = calc.get('fields', [])
            cfg['formula'] = calc.get('formula', '')
        configs.append(cfg)
        # 同时收集 calc_configs 中的非 increment 配置（increment 由独立回填处理）
        for extra in (t.get('calc_configs') or []):
            if extra.get('type') != 'increment' and extra.get('result_name'):
                configs.append({
                    'type_id': t.get('id'),
                    'type_name': t.get('name'),
                    'category': t.get('category', ''),
                    'calc_type': extra.get('type'),
                    'result_name': extra.get('result_name'),
                    'show_chart': extra.get('show_chart', False),
                    'max_value': extra.get('max_value')
                })
    return jsonify(configs)


@main.route('/api/dashboard-types/categories', methods=['GET'])
def api_dashboard_types_categories():
    """获取所有仪表盘分类（合并 JSON 配置 + 数据库中实际使用的分类）"""
    types_list = _load_dashboard_types()
    cats = set(t.get('category', '') for t in types_list if t.get('category', '').strip())
    session = SessionLocal()
    try:
        for row in session.query(InspectionObject.device_type).filter(
            InspectionObject.status == 'active',
            InspectionObject.device_type.isnot(None),
            InspectionObject.device_type != ''
        ).distinct():
            if row[0]:
                cats.add(row[0])
    finally:
        session.close()
    return jsonify(sorted(cats))


@main.route('/api/dashboard-types', methods=['POST'])
def api_dashboard_types_add():
    """添加仪表盘类型"""
    data = request.get_json()
    if not data or not data.get('name', '').strip():
        return jsonify({'error': '类型名称不能为空'}), 400

    types_list = _load_dashboard_types()
    new_id = data.get('id', '').strip() or re.sub(r'[^a-z0-9_]', '', data['name'].lower().replace(' ', '_')) or f'type_{int(__import__("time").time() * 1000)}'

    if not new_id:
        return jsonify({'error': '类型ID不能为空'}), 400

    if any(t['id'] == new_id for t in types_list):
        return jsonify({'error': f'ID "{new_id}" 已存在'}), 409

    new_labels = data.get('labels', {})
    # 清理 labels 中的计算类指标
    calc_result_names = set()
    new_calc = data.get('calc_config')
    if new_calc and new_calc.get('result_name'):
        calc_result_names.add(new_calc['result_name'])
    for c in (data.get('calc_configs') or []):
        if c.get('result_name'):
            calc_result_names.add(c['result_name'])
    if calc_result_names and isinstance(new_labels, dict):
        new_labels = {k: v for k, v in new_labels.items() if v not in calc_result_names}

    new_type = {
        'id': new_id,
        'name': data['name'].strip(),
        'category': data.get('category', '').strip(),
        'description': data.get('description', '').strip(),
        'detect_keywords': data.get('detect_keywords', []),
        'labels': new_labels,
        'formulas': data.get('formulas', {}),
        'calc_config': data.get('calc_config', None),
        'calc_configs': data.get('calc_configs', []),
        'extra_labels': data.get('extra_labels', []),
        'result_rules': data.get('result_rules', {}),
        'number_before_label': data.get('number_before_label', False),
        'skip_location_match': data.get('skip_location_match', False),
    }
    types_list.append(new_type)
    _save_dashboard_types(types_list)
    # 双向同步 + 虚拟指标同步
    try:
        session = SessionLocal()
        _sync_virtual_metrics_for_dtype(new_type, session)
        obj = session.query(InspectionObject).filter_by(name=new_type['name'], status='active').first()
        if not obj and new_id.startswith('obj_'):
            try:
                obj = session.query(InspectionObject).get(int(new_id[4:]))
            except (ValueError, IndexError):
                pass
        if obj:
            _sync_bidirectional(obj.id, new_id, session)
        session.commit()
        session.close()
    except Exception as e:
        logger.error(f'  SYNC: 新增类型同步失败: {e}')
    return jsonify({'ok': True, 'type': new_type})


@main.route('/api/dashboard-types/<type_id>', methods=['PUT'])
def api_dashboard_types_update(type_id):
    """更新仪表盘类型"""
    data = request.get_json()
    if not data:
        return jsonify({'error': '请提供数据'}), 400

    types_list = _load_dashboard_types()
    for i, t in enumerate(types_list):
        if t['id'] == type_id:
            new_labels = data.get('labels', t['labels'])
            # 清理 labels 中的计算类指标（result_name 不应出现在标签映射中）
            calc_result_names = set()
            new_calc = data.get('calc_config', t.get('calc_config'))
            if new_calc and new_calc.get('result_name'):
                calc_result_names.add(new_calc['result_name'])
            for c in (data.get('calc_configs', t.get('calc_configs', [])) or []):
                if c.get('result_name'):
                    calc_result_names.add(c['result_name'])
            if calc_result_names and isinstance(new_labels, dict):
                new_labels = {k: v for k, v in new_labels.items() if v not in calc_result_names}
            types_list[i].update({
                'name': data.get('name', t['name']),
                'category': data.get('category', t.get('category', '')),
                'description': data.get('description', t['description']),
                'detect_keywords': data.get('detect_keywords', t['detect_keywords']),
                'labels': new_labels,
                'formulas': data.get('formulas', t.get('formulas', {})),
                'calc_config': data.get('calc_config', t.get('calc_config', None)),
                'calc_configs': data.get('calc_configs', t.get('calc_configs', [])),
                'extra_labels': data.get('extra_labels', t.get('extra_labels', [])),
                'result_rules': data.get('result_rules', t.get('result_rules', {})),
                'number_before_label': data.get('number_before_label', t.get('number_before_label', False)),
                'skip_location_match': data.get('skip_location_match', t.get('skip_location_match', False)),
            })
            _save_dashboard_types(types_list)
            # 双向同步 + 虚拟指标同步
            try:
                session = SessionLocal()
                # 检查 calc_config / calc_configs 中 result_name 是否改名，同步更新历史记录
                old_calc = t.get('calc_config')
                new_calc = types_list[i].get('calc_config')
                old_configs = t.get('calc_configs') or []
                new_configs = types_list[i].get('calc_configs') or []
                rename_pairs = []
                if old_calc and new_calc and old_calc.get('result_name') and new_calc.get('result_name'):
                    if old_calc['result_name'] != new_calc['result_name']:
                        rename_pairs.append((old_calc['result_name'], new_calc['result_name']))
                for old_c, new_c in zip(old_configs, new_configs):
                    if old_c.get('type') == 'increment' and new_c.get('type') == 'increment':
                        if old_c.get('result_name') and new_c.get('result_name') and old_c['result_name'] != new_c['result_name']:
                            rename_pairs.append((old_c['result_name'], new_c['result_name']))
                if rename_pairs:
                    obj = session.query(InspectionObject).filter_by(name=types_list[i]['name'], status='active').first()
                    if not obj and type_id.startswith('obj_'):
                        try:
                            obj = session.query(InspectionObject).get(int(type_id[4:]))
                        except (ValueError, IndexError):
                            pass
                    if obj:
                        records = session.query(InspectionRecord).filter_by(object_id=obj.id).all()
                        for r in records:
                            m = json.loads(r.metrics) if r.metrics else {}
                            changed = False
                            for old_key, new_key in rename_pairs:
                                if old_key in m:
                                    m[new_key] = m.pop(old_key)
                                    changed = True
                            if changed:
                                r.metrics = json.dumps(m, ensure_ascii=False)
                        # 重命名 ObjectMetric（先删旧再建新，避免重复）
                        for old_key, new_key in rename_pairs:
                            old_metric = session.query(ObjectMetric).filter_by(object_id=obj.id, key=old_key).first()
                            if old_metric:
                                # 检查 new_key 是否已存在，存在则先删旧的
                                existing_new = session.query(ObjectMetric).filter_by(object_id=obj.id, key=new_key).first()
                                if existing_new and existing_new.id != old_metric.id:
                                    session.delete(existing_new)
                                old_metric.key = new_key
                                old_metric.name = new_key
                _sync_virtual_metrics_for_dtype(types_list[i], session)
                obj = session.query(InspectionObject).filter_by(name=types_list[i]['name'], status='active').first()
                if not obj and type_id.startswith('obj_'):
                    try:
                        obj = session.query(InspectionObject).get(int(type_id[4:]))
                    except (ValueError, IndexError):
                        pass
                if obj:
                    # 清理改名后残留的旧指标
                    for old_key, new_key in rename_pairs:
                        stale = session.query(ObjectMetric).filter_by(object_id=obj.id, key=old_key).first()
                        if stale:
                            session.delete(stale)
                    _sync_bidirectional(obj.id, type_id, session)
                session.commit()
                session.close()
            except Exception as e:
                logger.error(f'  SYNC: 更新类型同步失败: {e}')
            return jsonify({'ok': True, 'type': types_list[i]})

    return jsonify({'error': f'类型 "{type_id}" 不存在'}), 404


@main.route('/api/dashboard-types/<type_id>', methods=['DELETE'])
def api_dashboard_types_delete(type_id):
    """删除仪表盘类型"""
    types_list = _load_dashboard_types()
    new_list = [t for t in types_list if t['id'] != type_id]
    if len(new_list) == len(types_list):
        return jsonify({'error': f'类型 "{type_id}" 不存在'}), 404
    _save_dashboard_types(new_list)
    return jsonify({'ok': True})


@main.route('/api/dashboard-types/sync', methods=['POST'])
def api_dashboard_types_sync():
    """从巡检对象同步指标到仪表盘类型"""
    session = SessionLocal()
    try:
        objects = session.query(InspectionObject).filter_by(status='active').all()
        types_list = _load_dashboard_types()
        updated = 0
        created = 0

        for obj in objects:
            metrics = session.query(ObjectMetric).filter_by(object_id=obj.id).all()
            if not metrics:
                continue

            labels = {}
            for m in metrics:
                labels[m.name] = m.key

            # 查找已存在的仪表盘类型（按 obj_X id 或 名称匹配）
            type_id = f'obj_{obj.id}'
            existing = next((t for t in types_list if t['id'] == type_id), None)
            if not existing:
                existing = next((t for t in types_list if t['name'] == obj.name), None)

            if existing:
                if existing.get('labels') != labels:
                    existing['labels'] = labels
                    existing['detect_keywords'] = list(set(existing.get('detect_keywords', []) + [obj.name]))
                    updated += 1
            else:
                types_list.append({
                    'id': type_id,
                    'name': obj.name,
                    'description': obj.description or f'{obj.name} 仪表盘',
                    'detect_keywords': [obj.name],
                    'labels': labels,
                    'extra_labels': [],
                    'result_rules': {},
                    'number_before_label': False,
                })
                created += 1

        _save_dashboard_types(types_list)
        return jsonify({'ok': True, 'created': created, 'updated': updated, 'total_objects': len(objects)})
    finally:
        session.close()


@main.route('/api/dashboard-types/<type_id>/sync', methods=['POST'])
def api_dashboard_type_sync_single(type_id):
    """从巡检对象同步指标到指定仪表盘类型"""
    session = SessionLocal()
    try:
        types_list = _load_dashboard_types()
        target = next((t for t in types_list if t['id'] == type_id), None)
        if not target:
            return jsonify({'error': f'类型 "{type_id}" 不存在'}), 404

        # 通过 id (obj_X) 或名称匹配巡检对象
        obj = None
        if type_id.startswith('obj_'):
            try:
                obj_id = int(type_id[4:])
                obj = session.query(InspectionObject).filter_by(id=obj_id, status='active').first()
            except (ValueError, IndexError):
                pass
        if not obj:
            obj = session.query(InspectionObject).filter_by(name=target['name'], status='active').first()
        if not obj:
            return jsonify({'error': '未找到对应的巡检对象，请先在对象管理中创建'}), 404

        metrics = session.query(ObjectMetric).filter_by(object_id=obj.id).all()
        if not metrics:
            return jsonify({'error': f'巡检对象 "{obj.name}" 没有配置指标'}), 400

        labels = {m.name: m.key for m in metrics}
        target['labels'] = labels
        target['detect_keywords'] = list(set(target.get('detect_keywords', []) + [obj.name]))
        _save_dashboard_types(types_list)
        return jsonify({'ok': True, 'updated_labels': labels})
    finally:
        session.close()


@main.route('/api/ocr/model-info', methods=['GET'])
def api_ocr_model_info():
    """获取当前 OCR 模型信息"""
    info = {
        'installed': ocr_engine is not None,
        'engine': 'RapidOCR (ONNX Runtime)',
        'version': None,
        'models': {},
    }
    if ocr_engine is not None:
        try:
            import rapidocr_onnxruntime
            info['version'] = getattr(rapidocr_onnxruntime, '__version__', None)
        except Exception:
            pass
        try:
            det_cls_rec = getattr(ocr_engine, 'det_cls_rec', None)
            if det_cls_rec:
                det_model, cls_model, rec_model = det_cls_rec
                info['models'] = {
                    'detection': getattr(det_model, 'model_path', None) or str(getattr(det_model, 'model', '')),
                    'classification': getattr(cls_model, 'model_path', None) or str(getattr(cls_model, 'model', '')),
                    'recognition': getattr(rec_model, 'model_path', None) or str(getattr(rec_model, 'model', '')),
                }
        except Exception:
            pass
        try:
            textdet = getattr(ocr_engine, 'textdet', None)
            if textdet:
                info['models']['detection'] = getattr(getattr(textdet, 'model', None), 'model_path', None) or info['models'].get('detection', '')
            textrec = getattr(ocr_engine, 'textrec', None)
            if textrec:
                info['models']['recognition'] = getattr(getattr(textrec, 'model', None), 'model_path', None) or info['models'].get('recognition', '')
        except Exception:
            pass
    return jsonify(info)


@main.route('/api/objects/quick-create', methods=['POST'])
def api_quick_create_object():
    """快速创建巡检对象（JSON API）"""
    data = request.get_json()
    if not data or not data.get('name', '').strip():
        return jsonify({'error': '名称不能为空'}), 400

    name = data['name'].strip()
    location = data.get('location', '').strip() or None
    device_type = data.get('device_type', '').strip() or None
    description = data.get('description', '').strip() or None
    project_url = data.get('project_url', '').strip() or None

    session = SessionLocal()
    try:
        query = session.query(InspectionObject).filter_by(name=name, status='active')
        if location:
            query = query.filter_by(location=location)
        existing = query.first()
        if existing:
            msg = f'对象 "{name}" 已存在'
            if location:
                msg += f'（位置: {location}）'
            return jsonify({'error': msg, 'id': existing.id}), 409

        obj = InspectionObject(name=name, location=location, device_type=device_type, description=description, project_url=project_url)
        session.add(obj)
        session.commit()

        # 创建对应的仪表盘类型
        type_id = f'obj_{obj.id}'
        _sync_dashboard_type_from_object(obj, [])

        # 双向同步指标
        if device_type:
            types_list = _load_dashboard_types()
            matched_type = next((t for t in types_list if t.get('category') == device_type or t.get('name') == device_type), None)
            if matched_type:
                _sync_bidirectional(obj.id, matched_type['id'], session)
                session.commit()

        return jsonify({'ok': True, 'id': obj.id, 'name': obj.name})
    except Exception as e:
        session.rollback()
        return jsonify({'error': f'创建失败: {str(e)}'}), 500
    finally:
        session.close()


@main.route('/api/objects/<int:object_id>/sync-metrics', methods=['POST'])
def api_object_sync_metrics(object_id):
    """从仪表盘类型同步指标到巡检对象"""
    data = request.get_json() or {}
    dashboard_type_name = data.get('dashboard_type_name', '').strip()
    if not dashboard_type_name:
        return jsonify({'error': '缺少仪表盘类型名称'}), 400

    types_list = _load_dashboard_types()
    dashboard_type = next((t for t in types_list if t['name'] == dashboard_type_name), None)
    if not dashboard_type:
        return jsonify({'error': f'未找到仪表盘类型 "{dashboard_type_name}"'}), 404

    labels = dashboard_type.get('labels', {})
    if not labels:
        return jsonify({'ok': True, 'created': 0})

    session = SessionLocal()
    try:
        obj = session.query(InspectionObject).get(object_id)
        if not obj:
            return jsonify({'error': '巡检对象不存在'}), 404

        existing_keys = {m.key for m in obj.metrics}
        created = 0
        for i, (label_name, label_key) in enumerate(labels.items()):
            if label_key not in existing_keys:
                m = ObjectMetric(object_id=object_id, key=label_key, name=label_name, sort_order=i)
                session.add(m)
                created += 1
        session.commit()
        return jsonify({'ok': True, 'created': created, 'total': len(labels)})
    except Exception as e:
        session.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        session.close()


@main.route('/api/ocr/test', methods=['POST'])
def api_ocr_test():
    """用当前配置测试 OCR 识别"""
    if not ocr_engine:
        return jsonify({'error': 'OCR 引擎未安装，请运行: pip install rapidocr-onnxruntime'}), 500

    data = request.get_json()
    if not data or 'image' not in data:
        return jsonify({'error': '请提供图片数据'}), 400

    image_data = data['image']
    if ',' in image_data:
        image_data = image_data.split(',', 1)[1]

    config = _load_ocr_config()
    ignore_top = config.get('ignore_top', 0)
    ignore_bottom = config.get('ignore_bottom', 0)

    try:
        image_bytes = base64.b64decode(image_data)
        img = Image.open(io.BytesIO(image_bytes))
        img_height = img.height
        img_width = img.width
        logger.info(f'  IMAGE: (test) 原始图片尺寸 {img_width}x{img_height}, 格式 {img.format}, 模式 {img.mode}')
        top_cutoff = int(img_height * ignore_top / 100) if ignore_top > 0 else 0
        bottom_cutoff = img_height - int(img_height * ignore_bottom / 100) if ignore_bottom > 0 else img_height

        with tempfile.NamedTemporaryFile(suffix='.png', delete=False) as tmp:
            img.save(tmp.name)
            tmp_path = tmp.name
        import time as _time
        t0 = _time.time()
        result, _ = ocr_engine(tmp_path)
        elapsed = _time.time() - t0
        logger.info(f'  IMAGE: (test) OCR耗时 {elapsed:.2f}s, 结果条数 {len(result) if result else 0}')
        Path(tmp_path).unlink(missing_ok=True)
    except Exception as e:
        return jsonify({'error': f'图片处理失败: {str(e)}'}), 400

    filtered_result = []
    for item in (result or []):
        if item and len(item) >= 2:
            bbox = item[0]
            if bbox and len(bbox) >= 4:
                y_center = (bbox[0][1] + bbox[2][1]) / 2
                if y_center < top_cutoff or y_center > bottom_cutoff:
                    continue
            filtered_result.append(item)

    raw_lines = [item[1] for item in filtered_result if item]
    return jsonify({'raw_lines': raw_lines, 'config': config})


@main.route('/api/backup/gallery')
def api_backup_gallery():
    """获取备份截图列表（按日期分组）"""
    backup_base = Path(__file__).parent / 'backup'
    if not backup_base.exists():
        return jsonify({'folders': []})
    folders = []
    for folder in sorted(backup_base.iterdir(), reverse=True):
        if not folder.is_dir() or not folder.name.isdigit():
            continue
        images = []
        for img in sorted(folder.iterdir()):
            if img.suffix.lower() in ('.png', '.jpg', '.jpeg', '.gif', '.webp'):
                images.append({
                    'name': img.name,
                    'size': img.stat().st_size,
                    'url': f'/backup/{folder.name}/{img.name}'
                })
        if images:
            date_str = f'{folder.name[:4]}-{folder.name[4:6]}-{folder.name[6:8]}'
            folders.append({'date': date_str, 'folder': folder.name, 'count': len(images), 'images': images})
    return jsonify({'folders': folders})


@main.route('/backup/<path:folder>/<filename>')
def serve_backup_image(folder, filename):
    """提供备份图片访问"""
    from flask import send_from_directory
    backup_dir = Path(__file__).parent / 'backup' / folder
    return send_from_directory(str(backup_dir), filename)


@main.route('/backup-gallery')
def backup_gallery_page():
    """截图预览页面"""
    return render_template('backup_gallery.html')


@main.route('/api/backup-today-path')
def api_backup_today_path():
    """获取今日备份文件夹绝对路径"""
    date_folder = datetime.now().strftime('%Y%m%d')
    backup_dir = Path(__file__).parent / 'backup' / date_folder
    backup_dir.mkdir(parents=True, exist_ok=True)
    return jsonify({'path': str(backup_dir.resolve())})


@main.route('/api/backup-screenshot', methods=['POST'])
def api_backup_screenshot():
    """自动备份截图到本地文件夹"""
    data = request.get_json()
    if not data or 'image' not in data:
        return jsonify({'error': '没有图片数据'}), 400

    import base64
    image_data = data['image']
    date_folder = data.get('date_folder', datetime.now().strftime('%Y%m%d'))
    filename = data.get('filename', f'screenshot_{datetime.now().strftime("%H%M%S")}.png')

    # 解析 base64 图片
    if ',' in image_data:
        image_data = image_data.split(',', 1)[1]
    img_bytes = base64.b64decode(image_data)

    # 保存到 backup/日期/ 文件夹
    backup_dir = Path(__file__).parent / 'backup' / date_folder
    backup_dir.mkdir(parents=True, exist_ok=True)
    file_path = backup_dir / filename
    file_path.write_bytes(img_bytes)

    logger.info(f'  BACKUP: 已保存截图 {file_path}')
    return jsonify({'ok': True, 'path': str(file_path)})


@main.route('/api/save', methods=['POST'])
def api_save():
    """保存 OCR 识别结果"""
    data = request.get_json()
    if not data or 'items' not in data:
        return jsonify({'error': '没有要保存的数据'}), 400

    items = data['items']

    # 保存截图备份
    screenshot = data.get('screenshot', '')
    if screenshot:
        try:
            import base64 as _b64
            img_data = screenshot
            if ',' in img_data:
                img_data = img_data.split(',', 1)[1]
            img_bytes = _b64.b64decode(img_data)
            date_folder = datetime.now().strftime('%Y%m%d')
            backup_name = f'screenshot_{datetime.now().strftime("%H%M%S")}.png'
            backup_dir = Path(__file__).parent / 'backup' / date_folder
            backup_dir.mkdir(parents=True, exist_ok=True)
            (backup_dir / backup_name).write_bytes(img_bytes)
            logger.info(f'  BACKUP: 已保存截图 {backup_dir / backup_name}')
        except Exception as e:
            logger.warning(f'  BACKUP: 截图备份失败 {e}')

    # 用 Flask session 获取当前登录人员（SQLAlchemy session 另用 db 变量避免冲突）
    _login_inspector_id = dict(session).get('inspector_id')
    db = SessionLocal()
    saved = 0
    skipped = 0
    skipped_duplicate = 0
    skipped_no_match = 0
    skipped_incomplete = 0
    skipped_reasons = []
    created = 0
    last_object_id = None
    try:
        all_objects = db.query(InspectionObject).filter_by(status='active').all()
        all_inspectors = db.query(Inspector).all()

        for item in items:
            object_name = item.get('point_name', '').strip() or item.get('object_name', '').strip()
            location = item.get('location', '').strip()
            result = item.get('result', '正常')
            status_detail = item.get('status_detail', '')
            notes = item.get('notes', '')
            inspector_name = item.get('inspector', '')
            timestamp_str = item.get('timestamp')
            is_dashboard = item.get('is_dashboard', False)
            dashboard_type_name = item.get('dashboard_type_name', '') if is_dashboard else ''
            dashboard_category = item.get('dashboard_category', '') if is_dashboard else ''
            front_point_id = item.get('point_id')

            # 匹配巡检对象
            obj = None
            skip_location = item.get('skip_location_match', False)

            # 优先使用前端已选择的 point_id（但 device_type 须匹配）
            if front_point_id:
                try:
                    obj = db.query(InspectionObject).get(int(front_point_id))
                    if obj and dashboard_category and obj.device_type and obj.device_type != dashboard_category:
                        logger.info(f'  SAVE: 前端选择对象 device_type={obj.device_type} 与类别 {dashboard_category} 不匹配，忽略')
                        obj = None
                except (ValueError, TypeError):
                    obj = None

            # 如果有明确的对象名称，先按名称匹配（仪表盘类型需 device_type 匹配）
            if not obj and object_name:
                obj = _match_object(object_name, all_objects, expected_type=dashboard_category)
            
            # 仪表盘格式：按位置匹配（跳过位置匹配时不执行）
            if not obj and is_dashboard and location and not skip_location:
                obj = _match_object_by_location(location, all_objects, expected_type=dashboard_category)
            
            # 如果还是没有匹配，创建新对象
            if not obj:
                if is_dashboard and location:
                    obj_name = dashboard_type_name or location
                    obj_location = object_name or location
                elif object_name:
                    obj_name = object_name
                    obj_location = location
                else:
                    skipped += 1
                    skipped_no_match += 1
                    continue
                
                obj = InspectionObject(
                    name=obj_name,
                    location=obj_location,
                    device_type=dashboard_category or '监控',
                    description=f'OCR自动创建 [{dashboard_type_name or "监控"}] - {status_detail}',
                    status='active'
                )
                db.add(obj)
                db.flush()
                all_objects.append(obj)
                created += 1

            # 匹配巡检人员
            inspector = None
            front_inspector_id = item.get('inspector_id')
            if front_inspector_id:
                try:
                    inspector = db.query(Inspector).get(int(front_inspector_id))
                except (ValueError, TypeError):
                    inspector = None
            if not inspector and inspector_name:
                inspector = _match_inspector(inspector_name, all_inspectors)
            # 使用当前登录人员
            if not inspector and _login_inspector_id:
                inspector = db.query(Inspector).get(_login_inspector_id)
            if not inspector and all_inspectors:
                inspector = all_inspectors[0]

            # 解析时间（优先使用图片创建时间）
            timestamp = None
            file_creation_time = item.get('file_creation_time')
            if file_creation_time:
                try:
                    timestamp = datetime.strptime(file_creation_time, '%Y-%m-%d %H:%M')
                except ValueError:
                    pass
            if not timestamp and timestamp_str:
                if isinstance(timestamp_str, str):
                    try:
                        timestamp = datetime.strptime(timestamp_str, '%Y-%m-%d %H:%M')
                    except ValueError:
                        timestamp = datetime.now()
                else:
                    timestamp = timestamp_str
            if not timestamp:
                timestamp = datetime.now()

            # 重复检测：同一对象 + 同一精确时间（同一张图不能重复，但同一天不同图允许）
            existing_record = db.query(InspectionRecord).filter(
                InspectionRecord.object_id == obj.id,
                InspectionRecord.timestamp == timestamp
            ).first()
            if existing_record:
                logger.info(f'  SAVE: 跳过重复记录 obj={obj.id} time={timestamp}')
                skipped += 1
                skipped_duplicate += 1
                continue

            # 根据指标阈值重新判断结果
            parsed_metrics = _parse_status_to_metrics(status_detail)

            # 获取仪表盘类型ID，用于构建标签映射
            dashboard_type_id = item.get('dashboard_type', '')

            # 指标完整性校验：仪表盘类型要求的标签必须全部识别到
            if dashboard_type_id and is_dashboard:
                dashboard_types_check = _load_dashboard_types()
                matched_dtype_check = next((dt for dt in dashboard_types_check if dt.get('id') == dashboard_type_id), None)
                if matched_dtype_check:
                    required_labels = set(matched_dtype_check.get('labels', {}).values())
                    calc_config = matched_dtype_check.get('calc_config') or {}
                    calc_result = calc_config.get('result_name', '')
                    # 去掉计算类指标和 extra_labels（如在线率，OCR 无法直接识别）
                    required_labels.discard(calc_result)
                    # 去掉增量类指标（由保存时自动计算，OCR 不识别）
                    for c in (matched_dtype_check.get('calc_configs') or []):
                        if c.get('type') == 'increment' and c.get('result_name'):
                            required_labels.discard(c['result_name'])
                    for extra in matched_dtype_check.get('extra_labels', []):
                        required_labels.discard(extra)
                    parsed_keys = set(parsed_metrics.keys())
                    missing = required_labels - parsed_keys
                    if missing:
                        logger.info(f'  SAVE: 跳过不完整记录 obj={object_name} 缺少标签 {missing} parsed={list(parsed_metrics.keys())}')
                        skipped += 1
                        skipped_incomplete += 1
                        skipped_reasons.append({'name': object_name, 'reason': f'缺少: {", ".join(missing)}'})
                        continue

            # 构建中文标签 -> 短key 的反向映射（用于匹配已有指标）
            label_to_key = {}
            if dashboard_type_id:
                dashboard_types = _load_dashboard_types()
                matched_dtype = next((dt for dt in dashboard_types if dt.get('id') == dashboard_type_id), None)
                if matched_dtype:
                    for cn_label, short_key in matched_dtype.get('labels', {}).items():
                        label_to_key[cn_label] = short_key

            # 自动创建指标配置（如果不存在）
            if parsed_metrics:
                existing_metrics = {m.key: m for m in db.query(ObjectMetric).filter_by(object_id=obj.id).all()}
                existing_names = {m.name: m for m in db.query(ObjectMetric).filter_by(object_id=obj.id).all()}
                
                # 获取 calc_config 用于设置 show_chart 和 max_value
                calc_config = matched_dtype.get('calc_config', {}) if matched_dtype else {}
                calc_result_name = calc_config.get('result_name', '') if calc_config else ''
                
                for key, val in parsed_metrics.items():
                    try:
                        fv = float(str(val).rstrip('%'))
                    except (ValueError, TypeError):
                        continue
                    # 用中文标签对应的短key去匹配已有指标
                    mapped_key = label_to_key.get(key, key)
                    match_key = mapped_key if mapped_key in existing_metrics else key
                    # 用中文名查找已有指标（避免重复）
                    cn_name = None
                    for cn, sk in label_to_key.items():
                        if sk == key or sk == mapped_key:
                            cn_name = cn
                            break
                    
                    # 判断是否为计算字段，获取其 show_chart 和 max_value
                    is_calc_field = (key == calc_result_name or mapped_key == calc_result_name)
                    calc_show_chart = calc_config.get('show_chart', True) if is_calc_field and calc_config else True
                    calc_max_value = calc_config.get('max_value', None) if is_calc_field and calc_config else None
                    
                    if match_key in existing_metrics:
                        m = existing_metrics[match_key]
                        # 如果是计算字段且配置了 max_value，使用配置值
                        if is_calc_field and calc_max_value is not None:
                            m.max_value = calc_max_value
                        # 更新 show_in_chart
                        if is_calc_field:
                            m.show_in_chart = calc_show_chart
                        if fv >= 10000 and not m.unit:
                            m.unit = 'w'
                    elif cn_name and cn_name in existing_names:
                        pass  # 已有同名指标，跳过
                    else:
                        unit = ''
                        if is_calc_field and calc_max_value is not None:
                            max_val = calc_max_value
                        else:
                            max_val = 100
                            if fv >= 10000:
                                unit = 'w'
                        db.add(ObjectMetric(
                            object_id=obj.id, key=mapped_key, name=cn_name or mapped_key, unit=unit,
                            max_value=max_val, show_in_chart=calc_show_chart
                        ))
                        logger.info(f'  SAVE: 自动创建指标配置 key={mapped_key} name={cn_name or mapped_key} max_value={max_val} show_chart={calc_show_chart}')

            threshold_result = _check_metrics_thresholds(parsed_metrics, obj.id, db)
            if threshold_result:
                result = threshold_result

            # delta 规则评估（日对比）
            if dashboard_type_id:
                dashboard_types = _load_dashboard_types()
                matched_dtype = next((dt for dt in dashboard_types if dt.get('id') == dashboard_type_id), None)
                if matched_dtype:
                    result_rules = matched_dtype.get('result_rules', {})
                    delta_result = _evaluate_delta_rules(parsed_metrics, obj.id, result_rules, db)
                    if delta_result:
                        result = delta_result

            record = InspectionRecord(
                point_id=obj.id,
                object_id=obj.id,
                inspector_id=inspector.id if inspector else None,
                result=result,
                status_detail=status_detail,
                metrics=json.dumps(parsed_metrics, ensure_ascii=False),
                notes=notes,
                timestamp=timestamp
            )
            db.add(record)
            db.flush()  # 分配 record.id，确保后续查询排除自身

            # 增量计算（与上一条记录对比）
            if matched_dtype:
                # 检查 calc_config 或 calc_configs 中的 increment 配置
                calc_configs_list = matched_dtype.get('calc_configs') or []
                single_config = matched_dtype.get('calc_config')
                if single_config and single_config.get('type') == 'increment':
                    calc_configs_list = [single_config] + calc_configs_list
                for calc_config in calc_configs_list:
                    if calc_config.get('type') != 'increment':
                        continue
                    source_field = calc_config.get('source_field', '')
                    result_name = calc_config.get('result_name', '')
                    unit = calc_config.get('unit', '')
                    if source_field and result_name:
                        curr_val = parsed_metrics.get(source_field)
                        if curr_val is not None:
                            try:
                                curr_num = float(str(curr_val).rstrip('%'))
                                # 查找上一条记录
                                prev_record = db.query(InspectionRecord).filter(
                                    InspectionRecord.object_id == obj.id,
                                    InspectionRecord.id != record.id
                                ).order_by(InspectionRecord.timestamp.desc()).first()
                                if prev_record and prev_record.metrics:
                                    prev_metrics = json.loads(prev_record.metrics) if isinstance(prev_record.metrics, str) else prev_record.metrics
                                    pv = prev_metrics.get(source_field)
                                    prev_num = float(str(pv).rstrip('%')) if pv is not None else 0
                                    increment = curr_num - prev_num
                                    # 按天数平滑：间隔>1天时除以天数
                                    gap_days = (record.timestamp - prev_record.timestamp).total_seconds() / 86400
                                    if gap_days > 1:
                                        increment = increment / gap_days
                                else:
                                    # 首条记录，无对比基准，增量为0
                                    increment = 0
                                increment = round(increment, 1)
                                parsed_metrics[result_name] = increment
                                record.metrics = json.dumps(parsed_metrics, ensure_ascii=False)
                                logger.info(f'  SAVE: 增量计算 {result_name} = {curr_num} - prev = {increment}')
                                # 自动创建增量指标配置
                                existing_metrics = {m.key: m for m in db.query(ObjectMetric).filter_by(object_id=obj.id).all()}
                                if result_name not in existing_metrics:
                                    db.add(ObjectMetric(
                                        object_id=obj.id, key=result_name, name=result_name, unit=unit,
                                        max_value=calc_config.get('max_value'), show_in_chart=calc_config.get('show_chart', True)
                                    ))
                                else:
                                    m = existing_metrics[result_name]
                                    if calc_config.get('max_value') is not None:
                                        m.max_value = calc_config['max_value']
                                    m.show_in_chart = calc_config.get('show_chart', True)
                            except (ValueError, TypeError) as e:
                                logger.info(f'  SAVE: 增量计算失败 {e}')

            # 计数模式：保存日列表记录
            list_items = item.get('_list_items', [])
            if list_items:
                today = date.today()
                raw_count = item.get('_raw_count', len(list_items))
                unique_count = len(list_items)
                
                # UPSERT：同一天同一对象更新
                existing_daily = db.query(DailyListRecord).filter_by(
                    object_id=obj.id, date=today, content_type='list'
                ).first()
                
                if existing_daily:
                    existing_daily.items = json.dumps(list_items, ensure_ascii=False)
                    existing_daily.count = unique_count
                    existing_daily.raw_count = raw_count
                    logger.info(f'  SAVE: 更新日列表记录 obj_id={obj.id} date={today} count={unique_count}')
                else:
                    daily_record = DailyListRecord(
                        object_id=obj.id,
                        date=today,
                        content_type='list',
                        items=json.dumps(list_items, ensure_ascii=False),
                        count=unique_count,
                        raw_count=raw_count
                    )
                    db.add(daily_record)
                    logger.info(f'  SAVE: 新增日列表记录 obj_id={obj.id} date={today} count={unique_count}')

            db.commit()
            saved += 1
            last_object_id = obj.id
    except Exception as e:
        db.rollback()
        return jsonify({'error': f'保存失败: {str(e)}'}), 500
    finally:
        db.close()

    return jsonify({'saved': saved, 'skipped': skipped, 'skipped_duplicate': skipped_duplicate, 'skipped_no_match': skipped_no_match, 'skipped_incomplete': skipped_incomplete, 'skipped_reasons': skipped_reasons, 'created': created, 'object_id': last_object_id})


def _match_object(name, all_objects, expected_type=None):
    """匹配巡检对象"""
    name_lower = name.lower()
    best = None
    best_score = 0
    expected_lower = expected_type.lower() if expected_type else None

    for obj in all_objects:
        if expected_lower and obj.device_type and obj.device_type.lower() != expected_lower:
            continue
        score = 0
        has_name_match = False
        obj_name_lower = (obj.name or '').lower()
        location_lower = (obj.location or '').lower()

        if obj_name_lower and (obj_name_lower in name_lower or name_lower in obj_name_lower):
            score += 20
            has_name_match = True
        if location_lower and location_lower in name_lower:
            score += 10
        if obj.device_type:
            dtype_lower = obj.device_type.lower()
            if dtype_lower in name_lower:
                score += 5

        if has_name_match and score > best_score:
            best_score = score
            best = obj

    return best


def _match_object_by_location(location, all_objects, expected_type=None):
    """按位置匹配巡检对象"""
    if not location:
        return None
    
    location_lower = location.lower()
    best = None
    best_score = 0
    expected_lower = expected_type.lower() if expected_type else None

    for obj in all_objects:
        if expected_lower and obj.device_type and obj.device_type.lower() != expected_lower:
            continue
        score = 0
        has_location_match = False
        obj_location_lower = (obj.location or '').lower()

        # 位置完全匹配
        if obj_location_lower and obj_location_lower == location_lower:
            score += 30
            has_location_match = True
        # 位置包含匹配
        elif obj_location_lower and location_lower in obj_location_lower:
            score += 20
            has_location_match = True
        elif obj_location_lower and obj_location_lower in location_lower:
            score += 15
            has_location_match = True

        if has_location_match and score > best_score:
            best_score = score
            best = obj

    return best


def _match_object_by_type(device_type, all_objects):
    """按设备类型匹配巡检对象（返回第一个匹配的）"""
    if not device_type:
        return None
    
    type_lower = device_type.lower()
    for obj in all_objects:
        if (obj.device_type or '').lower() == type_lower:
            return obj
    return None


def _match_inspector(name, all_inspectors):
    """匹配巡检人员"""
    name_lower = name.lower()
    for inspector in all_inspectors:
        if (inspector.name or '').lower() == name_lower:
            return inspector
    return None


@main.route('/api/export/json')
def api_export_json():
    """导出 JSON 数据 API"""
    session = SessionLocal()
    try:
        objects = session.query(InspectionObject).filter_by(status='active').all()
        data = []
        for obj in objects:
            records = (
                session.query(InspectionRecord)
                .filter_by(object_id=obj.id)
                .order_by(InspectionRecord.timestamp.desc())
                .all()
            )
            records_data = [
                {
                    'timestamp': r.timestamp.strftime('%Y-%m-%d %H:%M'),
                    'result': r.result,
                    'status_detail': r.status_detail,
                    'notes': r.notes,
                    'inspector': r.inspector.name if r.inspector else '未知',
                }
                for r in records
            ]
            data.append({
                'name': obj.name,
                'location': obj.location,
                'device_type': obj.device_type,
                'description': obj.description,
                'records': records_data,
            })

        return Response(
            json.dumps(data, ensure_ascii=False, indent=2),
            mimetype='application/json',
            headers={'Content-Disposition': 'attachment; filename="inspection_data.json"'}
        )
    finally:
        session.close()